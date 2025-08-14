import cv2
import numpy as np
from pathlib import Path
import ttkbootstrap as ttk
from ttkbootstrap.constants import (LEFT, BOTH, DISABLED, NORMAL, X, Y, BOTTOM, RIGHT, HORIZONTAL, VERTICAL, NW, CENTER)
import tkinter as tk
from tkinter import (Canvas, filedialog, messagebox, simpledialog, Toplevel, StringVar, Text,
                     colorchooser, DoubleVar)
from tkinter.ttk import Combobox
from PIL import Image, ImageTk
from datetime import datetime
import os
import time

# Importa módulos do sistema de banco de dados
try:
    # Quando importado como módulo
    from database_manager import DatabaseManager
    from model_selector import ModelSelectorDialog, SaveModelDialog
    from dialogs import EditSlotDialog, SystemConfigDialog
    from training_dialog import SlotTrainingDialog
    from utils import load_style_config, save_style_config, apply_style_config, get_style_config_path, get_color, get_colors_group, get_font
    from ml_classifier import MLSlotClassifier
    from camera_manager import (
        detect_cameras,
        get_cached_camera,
        release_cached_camera,
        cleanup_unused_cameras,
        release_all_cached_cameras,
        capture_image_from_camera,
    )
    from image_utils import cv2_to_tk
except ImportError:
    # Quando executado diretamente
    try:
        from database_manager import DatabaseManager
        from model_selector import ModelSelectorDialog, SaveModelDialog
        from dialogs import EditSlotDialog, SystemConfigDialog
        from training_dialog import SlotTrainingDialog
        from utils import load_style_config, save_style_config, apply_style_config, get_style_config_path, get_color, get_colors_group, get_font
        from ml_classifier import MLSlotClassifier
        from camera_manager import (
            detect_cameras,
            get_cached_camera,
            release_cached_camera,
            cleanup_unused_cameras,
            release_all_cached_cameras,
            capture_image_from_camera,
        )
        from image_utils import cv2_to_tk
    except ImportError:
        # Quando executado a partir do diretório raiz
        from modulos.database_manager import DatabaseManager
        from modulos.model_selector import ModelSelectorDialog, SaveModelDialog
        from modulos.dialogs import EditSlotDialog, SystemConfigDialog
        from modulos.training_dialog import SlotTrainingDialog
        from modulos.utils import load_style_config, save_style_config, apply_style_config, get_style_config_path, get_color, get_colors_group, get_font
        from modulos.ml_classifier import MLSlotClassifier
        from modulos.camera_manager import (
            detect_cameras,
            get_cached_camera,
            release_cached_camera,
            cleanup_unused_cameras,
            release_all_cached_cameras,
            capture_image_from_camera,
        )
        from modulos.image_utils import cv2_to_tk

# Importar funções do módulo inspection para evitar duplicação
try:
    from inspection import find_image_transform, check_slot
except ImportError:
    try:
        from modulos.inspection import find_image_transform, check_slot
    except ImportError:
        # Fallback: definir funções vazias se não conseguir importar
        def find_image_transform(*args, **kwargs):
            raise ImportError("Módulo inspection não disponível")
        def check_slot(*args, **kwargs):
            raise ImportError("Módulo inspection não disponível")

# ---------- parâmetros globais ------------------------------------------------
# Caminho para a pasta de modelos na raiz do projeto
# Usa caminhos relativos para permitir portabilidade
try:
    from paths import (
        get_project_root,
        get_model_dir,
        get_template_dir,
        get_model_template_dir,
    )
except Exception:
    try:
        from paths import (
            get_project_root,
            get_model_dir,
            get_template_dir,
            get_model_template_dir,
        )
    except Exception:
        from modulos.paths import (
            get_project_root,
            get_model_dir,
            get_template_dir,
            get_model_template_dir,
        )

# Define diretórios globais
MODEL_DIR = get_model_dir()
TEMPLATE_DIR = get_template_dir()

# Limiares de inspeção
THR_CORR = 0.1  # Limiar para template matching (clips)
MIN_PX = 10      # Contagem mínima de pixels para template matching (clips)

# Parâmetros do Canvas e Preview
PREVIEW_W = 1200  # Largura máxima do canvas para exibição inicial (aumentada)
PREVIEW_H = 900  # Altura máxima do canvas para exibição inicial (aumentada)

# Parâmetros ORB para registro de imagem
ORB_FEATURES = 5000
ORB_SCALE_FACTOR = 1.2
ORB_N_LEVELS = 8

# Cores são agora carregadas do arquivo de configuração centralizado
# Veja config/style_config.json para personalizar as cores

# Caminho para o arquivo de configurações de estilo
STYLE_CONFIG_PATH = get_style_config_path()


# ---------- utilidades --------------------------------------------------------




# Inicialização otimizada do detector ORB
try:
    # Configurações otimizadas para melhor performance
    orb = cv2.ORB_create(
        nfeatures=ORB_FEATURES,
        scaleFactor=ORB_SCALE_FACTOR,
        nlevels=ORB_N_LEVELS,
        edgeThreshold=31,  # Reduz detecção em bordas para melhor performance
        firstLevel=0,      # Nível inicial da pirâmide
        WTA_K=2,          # Número de pontos para comparação
        scoreType=cv2.ORB_HARRIS_SCORE,  # Usa Harris score para melhor qualidade
        patchSize=31      # Tamanho do patch para descritores
    )
    print("Detector ORB inicializado com sucesso (configuração otimizada).")
except Exception as e:
    print(f"Erro ao inicializar ORB: {e}. O registro de imagem não funcionará.")
    orb = None

# Cache para descritores de imagem de referência (otimização)
_ref_image_cache = {
    'image_hash': None,
    'keypoints': None,
    'descriptors': None,
    'gray_image': None
}


# Função find_image_transform removida - agora importada do módulo inspection.py


def transform_rectangle(rect, M, img_shape):
    """
    Transforma um retângulo usando uma matriz de homografia.
    rect: (x, y, w, h)
    M: matriz de homografia 3x3
    img_shape: (height, width) da imagem de destino
    Retorna: (x, y, w, h) transformado ou None se inválido
    """
    if M is None:
        return None
    
    x, y, w, h = rect
    
    # Define os 4 cantos do retângulo
    corners = np.float32([
        [x, y],
        [x + w, y],
        [x + w, y + h],
        [x, y + h]
    ]).reshape(-1, 1, 2)
    
    try:
        # Transforma os cantos
        transformed_corners = cv2.perspectiveTransform(corners, M)
        
        # Calcula o bounding box dos cantos transformados
        x_coords = transformed_corners[:, 0, 0]
        y_coords = transformed_corners[:, 0, 1]
        
        x_min, x_max = np.min(x_coords), np.max(x_coords)
        y_min, y_max = np.min(y_coords), np.max(y_coords)
        
        # Garante que está dentro dos limites da imagem
        img_h, img_w = img_shape[:2]
        x_min = max(0, int(x_min))
        y_min = max(0, int(y_min))
        x_max = min(img_w, int(x_max))
        y_max = min(img_h, int(y_max))
        
        new_w = x_max - x_min
        new_h = y_max - y_min
        
        if new_w <= 0 or new_h <= 0:
            print(f"Retângulo transformado inválido: ({x_min}, {y_min}, {new_w}, {new_h})")
            return None
        
        return (x_min, y_min, new_w, new_h)
        
    except Exception as e:
        print(f"Erro ao transformar retângulo: {e}")
        return None


def check_slot(img_test, slot_data, M):
    """
    Verifica um slot na imagem de teste.
    Retorna: (passou, correlation, pixels, corners, bbox, log_msgs)
    """
    log_msgs = []
    corners = None
    bbox = [0, 0, 0, 0]
    
    try:
        slot_type = slot_data.get('tipo', 'clip')
        x, y, w, h = slot_data['x'], slot_data['y'], slot_data['w'], slot_data['h']
        
        # Calcula os cantos originais do slot
        original_corners = [(x, y), (x+w, y), (x+w, y+h), (x, y+h)]
        
        # Transforma o retângulo se temos matriz de homografia
        # Respeita configuração de alinhamento por slot
        if slot_data.get('use_alignment', True) and M is not None:
            # Transforma os cantos usando a matriz de homografia
            corners_array = np.array(original_corners, dtype=np.float32).reshape(-1, 1, 2)
            transformed_corners = cv2.perspectiveTransform(corners_array, M)
            corners = [(int(pt[0][0]), int(pt[0][1])) for pt in transformed_corners]
            
            # Calcula bounding box dos cantos transformados
            x_coords = [pt[0] for pt in corners]
            y_coords = [pt[1] for pt in corners]
            x, y = max(0, min(x_coords)), max(0, min(y_coords))
            w = min(img_test.shape[1] - x, max(x_coords) - x)
            h = min(img_test.shape[0] - y, max(y_coords) - y)
            
            log_msgs.append(f"Slot transformado para ({x}, {y}, {w}, {h})")
        else:
            corners = original_corners
            log_msgs.append("Usando coordenadas originais (sem transformação)")
        
        bbox = [x, y, w, h]
        
        # Verifica se a ROI está dentro dos limites da imagem
        if x < 0 or y < 0 or x + w > img_test.shape[1] or y + h > img_test.shape[0]:
            log_msgs.append(f"ROI fora dos limites da imagem: ({x}, {y}, {w}, {h})")
            return False, 0.0, 0, corners, bbox, log_msgs
        
        # Extrai ROI
        roi = img_test[y:y+h, x:x+w]
        if roi.size == 0:
            log_msgs.append("ROI vazia")
            return False, 0.0, 0, corners, bbox, log_msgs
        
        if slot_type == 'clip':
            # Verifica se deve usar Machine Learning
            if slot_data.get('use_ml', False) and slot_data.get('ml_model_path'):
                try:
                    from ml_classifier import MLSlotClassifier
                    
                    # Carrega o modelo ML
                    ml_classifier = MLSlotClassifier()
                    ml_classifier.load_model(slot_data['ml_model_path'])
                    
                    # Faz a predição usando ML
                    prediction, confidence = ml_classifier.predict(roi)
                    
                    # Converte predição para resultado booleano
                    is_ok = prediction == 1  # 1 = OK, 0 = NG
                    
                    log_msgs.append(f"ML: Predição={prediction} ({'OK' if is_ok else 'NG'}), Confiança={confidence:.3f}")
                    return is_ok, confidence, 0, corners, bbox, log_msgs
                    
                except Exception as ml_error:
                    log_msgs.append(f"Erro no ML, usando método tradicional: {str(ml_error)}")
                    # Continua com método tradicional em caso de erro
            
            # Verifica método de detecção
            detection_method = slot_data.get('detection_method', 'template_matching')
            
            if detection_method == 'histogram_analysis':
                # === ANÁLISE POR HISTOGRAMA ===
                try:
                    # Calcula histograma da ROI em HSV
                    roi_hsv = cv2.cvtColor(roi, cv2.COLOR_BGR2HSV)
                    
                    # Parâmetros do histograma
                    h_bins = 50
                    s_bins = 60
                    hist_range = [0, 180, 0, 256]  # H: 0-179, S: 0-255
                    
                    # Calcula histograma 2D (H-S)
                    hist = cv2.calcHist([roi_hsv], [0, 1], None, [h_bins, s_bins], hist_range)
                    
                    # Normaliza histograma
                    cv2.normalize(hist, hist, 0, 1, cv2.NORM_MINMAX)
                    
                    # Calcula métricas do histograma
                    np.sum(hist)
                    np.mean(hist)
                    hist_std = np.std(hist)
                    hist_max = np.max(hist)
                    
                    # Calcula entropia do histograma
                    hist_flat = hist.flatten()
                    hist_flat = hist_flat[hist_flat > 0]  # Remove zeros
                    entropy = -np.sum(hist_flat * np.log2(hist_flat + 1e-10))
                    
                    # Score baseado em múltiplas métricas
                    # Combina entropia (diversidade de cores) e distribuição
                    entropy_score = min(entropy / 10.0, 1.0)  # Normaliza entropia
                    distribution_score = min(hist_std * 10, 1.0)  # Penaliza distribuições muito uniformes
                    intensity_score = min(hist_max * 2, 1.0)  # Considera picos de intensidade
                    
                    # Score final combinado
                    histogram_score = (entropy_score * 0.5 + distribution_score * 0.3 + intensity_score * 0.2)
                    
                    # Usa limiar personalizado do slot ou padrão
                    if 'correlation_threshold' in slot_data:
                        threshold = slot_data.get('correlation_threshold', 0.3)
                        threshold_source = "correlation_threshold"
                    else:
                        threshold = slot_data.get('detection_threshold', 30.0) / 100.0  # Converte % para decimal
                        threshold_source = "detection_threshold"
                    
                    # Usa somente correlação como critério final
                    passou = histogram_score >= threshold
                    
                    log_msgs.append(f"Histograma: {histogram_score:.3f} (limiar: {threshold:.2f} [{threshold_source}], entropia: {entropy:.2f}, std: {hist_std:.3f})")
                    return passou, histogram_score, 0, corners, bbox, log_msgs
                    
                except Exception as e:
                    log_msgs.append(f"Erro na análise por histograma: {str(e)}")
                    return False, 0.0, 0, corners, bbox, log_msgs
            
            elif detection_method == 'contour_analysis':
                # === ANÁLISE POR CONTORNO ===
                try:
                    # Converte para escala de cinza
                    roi_gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
                    
                    # Aplica blur para reduzir ruído
                    roi_blur = cv2.GaussianBlur(roi_gray, (5, 5), 0)
                    
                    # Detecta bordas com Canny
                    edges = cv2.Canny(roi_blur, 50, 150)
                    
                    # Encontra contornos
                    contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                    
                    # Se não encontrou contornos, retorna falha
                    if not contours:
                        log_msgs.append("Nenhum contorno encontrado")
                        return False, 0.0, 0, corners, bbox, log_msgs
                    
                    # Calcula área total da ROI
                    roi_area = roi.shape[0] * roi.shape[1]
                    
                    # Calcula área total dos contornos
                    contour_area = sum(cv2.contourArea(cnt) for cnt in contours)
                    
                    # Calcula número de contornos
                    num_contours = len(contours)
                    
                    # Calcula perímetro total
                    total_perimeter = sum(cv2.arcLength(cnt, True) for cnt in contours)
                    
                    # Calcula complexidade média dos contornos (razão perímetro/área)
                    complexity = total_perimeter / (contour_area + 1e-10)
                    
                    # Normaliza métricas
                    area_ratio = min(contour_area / roi_area, 1.0)  # Razão entre área de contornos e área total
                    contour_count_score = min(num_contours / 10.0, 1.0)  # Normaliza contagem de contornos
                    complexity_score = min(1.0, 1.0 / (complexity + 0.1))  # Inverte para que menor complexidade = maior score
                    
                    # Score final combinado
                    contour_score = (area_ratio * 0.4 + contour_count_score * 0.3 + complexity_score * 0.3)
                    
                    # Usa limiar personalizado do slot ou padrão
                    threshold = slot_data.get('detection_threshold', 0.5)
                    
                    # Usa somente correlação como critério final
                    passou = contour_score >= threshold
                    
                    log_msgs.append(f"Contorno: {contour_score:.3f} (limiar: {threshold:.2f}, contornos: {num_contours}, área: {area_ratio:.2f}, complexidade: {complexity:.2f})")
                    return passou, contour_score, 0, corners, bbox, log_msgs
                    
                except Exception as e:
                    log_msgs.append(f"Erro na análise por contorno: {str(e)}")
                    return False, 0.0, 0, corners, bbox, log_msgs
            
            elif detection_method == 'image_comparison':
                # === COMPARAÇÃO DIRETA DE IMAGEM ===
                try:
                    template_path = slot_data.get('template_path')
                    if not template_path or not Path(template_path).exists():
                        log_msgs.append("Template não encontrado para comparação de imagem")
                        return False, 0.0, 0, corners, bbox, log_msgs
                    
                    # Carrega o template
                    template = cv2.imread(str(template_path))
                    if template is None:
                        log_msgs.append("Erro ao carregar template para comparação de imagem")
                        return False, 0.0, 0, corners, bbox, log_msgs
                    
                    # Redimensiona o template para o tamanho da ROI
                    template_resized = cv2.resize(template, (roi.shape[1], roi.shape[0]))
                    
                    # Converte para escala de cinza
                    roi_gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY)
                    template_gray = cv2.cvtColor(template_resized, cv2.COLOR_BGR2GRAY)
                    
                    # Calcula SSIM (Structural Similarity Index)
                    from skimage.metrics import structural_similarity as ssim
                    try:
                        ssim_score, _ = ssim(roi_gray, template_gray, full=True)
                    except ImportError:
                        # Fallback se skimage não estiver disponível
                        # Calcula MSE (Mean Squared Error) e converte para similaridade
                        mse = np.mean((roi_gray.astype("float") - template_gray.astype("float")) ** 2)
                        ssim_score = 1 - (mse / 255**2)  # Normaliza para [0,1] onde 1 é perfeito
                    
                    # Calcula diferença absoluta
                    diff = cv2.absdiff(roi_gray, template_gray)
                    diff_score = 1.0 - (np.mean(diff) / 255.0)  # Normaliza para [0,1] onde 1 é perfeito
                    
                    # Calcula histogramas e compara
                    hist_roi = cv2.calcHist([roi_gray], [0], None, [256], [0, 256])
                    hist_template = cv2.calcHist([template_gray], [0], None, [256], [0, 256])
                    cv2.normalize(hist_roi, hist_roi, 0, 1, cv2.NORM_MINMAX)
                    cv2.normalize(hist_template, hist_template, 0, 1, cv2.NORM_MINMAX)
                    hist_score = cv2.compareHist(hist_roi, hist_template, cv2.HISTCMP_CORREL)
                    
                    # Score final combinado
                    comparison_score = (ssim_score * 0.5 + diff_score * 0.3 + hist_score * 0.2)
                    
                    # Usa limiar personalizado do slot ou padrão
                    threshold = slot_data.get('detection_threshold', 0.7)
                    
                    # Usa somente correlação como critério final
                    passou = comparison_score >= threshold
                    
                    log_msgs.append(f"Comparação: {comparison_score:.3f} (limiar: {threshold:.2f}, SSIM: {ssim_score:.2f}, Diff: {diff_score:.2f}, Hist: {hist_score:.2f})")
                    return passou, comparison_score, 0, corners, bbox, log_msgs
                    
                except Exception as e:
                    log_msgs.append(f"Erro na comparação de imagem: {str(e)}")
                    return False, 0.0, 0, corners, bbox, log_msgs
            
            else:  # template_matching (método padrão)
                # === TEMPLATE MATCHING PARA CLIPS ===
                template_path = slot_data.get('template_path')
                if not template_path or not Path(template_path).exists():
                    log_msgs.append("Template não encontrado")
                    return False, 0.0, 0, corners, bbox, log_msgs
                
                template = cv2.imread(str(template_path))
                if template is None:
                    log_msgs.append("Erro ao carregar template")
                    return False, 0.0, 0, corners, bbox, log_msgs
                
                # === TEMPLATE MATCHING OTIMIZADO ===
                slot_data.get('correlation_threshold', 0.7)
                template_method_str = slot_data.get('template_method', 'TM_CCOEFF_NORMED')
                scale_tolerance = slot_data.get('scale_tolerance', 10.0) / 100.0
                
                # Mapeamento otimizado de métodos
                method_map = {
                    'TM_CCOEFF_NORMED': cv2.TM_CCOEFF_NORMED,
                    'TM_CCORR_NORMED': cv2.TM_CCORR_NORMED,
                    'TM_SQDIFF_NORMED': cv2.TM_SQDIFF_NORMED
                }
                template_method = method_map.get(template_method_str, cv2.TM_CCOEFF_NORMED)
                
                max_val = 0.0
                best_scale = 1.0
                
                # Otimização: reduz número de escalas testadas
                if scale_tolerance > 0:
                    # Testa apenas 3 escalas para melhor performance
                    scales = [1.0 - scale_tolerance, 1.0, 1.0 + scale_tolerance]
                else:
                    scales = [1.0]  # Apenas escala original
                
                # Pré-converte template para escala de cinza se necessário (otimização)
                template_gray = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY) if len(template.shape) == 3 else template
                roi_gray = cv2.cvtColor(roi, cv2.COLOR_BGR2GRAY) if len(roi.shape) == 3 else roi
            
            for scale in scales:
                # Calcula dimensões da escala
                scaled_w = int(template_gray.shape[1] * scale)
                scaled_h = int(template_gray.shape[0] * scale)
                
                # Validação de dimensões otimizada
                if (scaled_w <= 0 or scaled_h <= 0 or 
                    scaled_w > roi_gray.shape[1] or scaled_h > roi_gray.shape[0]):
                    continue
                
                # Redimensiona template (usa INTER_AREA para redução, INTER_LINEAR para ampliação)
                interpolation = cv2.INTER_AREA if scale < 1.0 else cv2.INTER_LINEAR
                scaled_template = cv2.resize(template_gray, (scaled_w, scaled_h), interpolation=interpolation)
                
                # Template matching otimizado (usa imagens em escala de cinza)
                result = cv2.matchTemplate(roi_gray, scaled_template, template_method)
                
                # Extrai valor de correlação
                if template_method == cv2.TM_SQDIFF_NORMED:
                    min_val, _, _, _ = cv2.minMaxLoc(result)
                    current_val = 1.0 - min_val  # Inverte para SQDIFF
                else:
                    _, current_val, _, _ = cv2.minMaxLoc(result)
                
                # Atualiza melhor resultado
                if current_val > max_val:
                    max_val = current_val
                    best_scale = scale
            
            # Usa limiar personalizado do slot ou padrão
            # Prioridade: correlation_threshold > detection_threshold > padrão global
            if 'correlation_threshold' in slot_data:
                threshold = slot_data.get('correlation_threshold', 0.1)
                threshold_source = "correlation_threshold"
            else:
                threshold = slot_data.get('detection_threshold', 70.0) / 100.0  # Converte % para decimal
                threshold_source = "detection_threshold"
            
            # Usa apenas correlação como critério final
            passou = max_val >= threshold
            
            log_msgs.append(f"Correlação: {max_val:.3f} (limiar: {threshold:.2f} [{threshold_source}], escala: {best_scale:.2f}, método: {template_method_str})")
            return passou, max_val, 0, corners, bbox, log_msgs
        
        else:  # fita - tipo removido, apenas clips são suportados
            log_msgs.append("Tipo 'fita' não é mais suportado - apenas template matching para 'clip'")
            return False, 0.0, 0, corners, bbox, log_msgs
    
    except Exception as e:
        log_msgs.append(f"Erro: {str(e)}")
        print(f"Erro em check_slot: {e}")
        return False, 0.0, 0, corners, bbox, log_msgs




class SystemConfigDialog(Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.title("⚙️ Configurações do Sistema")
        self.geometry("550x750")  # Aumentado para acomodar todas as configurações
        self.resizable(True, True)  # Permitir redimensionamento para melhor visualização
        self.transient(parent)
        self.grab_set()
        
        # Importa o módulo colorchooser para seleção de cores
        
        # Carrega as configurações de estilo atuais
        self.style_config = load_style_config()
        
        self.result = False
        self.center_window()
        self.setup_ui()
    
    def center_window(self):
        self.update_idletasks()
        width = 550
        height = 750
        x = (self.winfo_screenwidth() // 2) - (width // 2)
        y = (self.winfo_screenheight() // 2) - (height // 2)
        self.geometry(f"{width}x{height}+{x}+{y}")
    
    def setup_ui(self):
        # Criar um canvas com scrollbar para acomodar todas as configurações
        container = ttk.Frame(self)
        container.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        # Criar canvas com tamanho adequado
        canvas = Canvas(container, width=530, height=700)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas, width=520)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        # Adicionar evento de rolagem com o mouse
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Configurar o layout
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Frame principal dentro do scrollable_frame
        main_frame = ttk.Frame(scrollable_frame)
        main_frame.pack(fill=BOTH, expand=True)
        
        # Configurações ORB
        orb_frame = ttk.LabelFrame(main_frame, text="Configurações ORB (Alinhamento de Imagem)", style='Section.TLabelframe')
        orb_frame.pack(fill=X, pady=(0, 10))
        
        ttk.Label(orb_frame, text="Número de Features:").pack(anchor="w", padx=5, pady=2)
        self.orb_features_var = ttk.IntVar(value=ORB_FEATURES)
        features_frame = ttk.Frame(orb_frame)
        features_frame.pack(fill=X, padx=5, pady=5)
        
        self.features_scale = ttk.Scale(features_frame, from_=1000, to=10000, variable=self.orb_features_var, orient=HORIZONTAL)
        self.features_scale.pack(side=LEFT, fill=X, expand=True)
        
        self.features_label = ttk.Label(features_frame, text=f"{self.orb_features_var.get()}", width=8)
        self.features_label.pack(side=RIGHT, padx=(5, 0))
        
        def update_features_label(val):
            self.features_label.config(text=f"{int(float(val))}")
        self.features_scale.config(command=update_features_label)
        
        ttk.Label(orb_frame, text="Fator de Escala:").pack(anchor="w", padx=5, pady=(10, 2))
        self.scale_factor_var = ttk.DoubleVar(value=ORB_SCALE_FACTOR)
        scale_frame = ttk.Frame(orb_frame)
        scale_frame.pack(fill=X, padx=5, pady=5)
        
        self.scale_scale = ttk.Scale(scale_frame, from_=1.1, to=2.0, variable=self.scale_factor_var, orient=HORIZONTAL)
        self.scale_scale.pack(side=LEFT, fill=X, expand=True)
        
        self.scale_label = ttk.Label(scale_frame, text=f"{self.scale_factor_var.get():.2f}", width=8)
        self.scale_label.pack(side=RIGHT, padx=(5, 0))
        
        def update_scale_label(val):
            self.scale_label.config(text=f"{float(val):.2f}")
        self.scale_scale.config(command=update_scale_label)
        
        ttk.Label(orb_frame, text="Número de Níveis:").pack(anchor="w", padx=5, pady=(10, 2))
        self.n_levels_var = ttk.IntVar(value=ORB_N_LEVELS)
        levels_spin = ttk.Spinbox(orb_frame, from_=4, to=16, textvariable=self.n_levels_var, width=10)
        levels_spin.pack(anchor="w", padx=5, pady=5)
        
        # Configurações de Canvas
        canvas_frame = ttk.LabelFrame(main_frame, text="Configurações de Visualização", style='Section.TLabelframe')
        canvas_frame.pack(fill=X, pady=(0, 10))
        
        ttk.Label(canvas_frame, text="Largura Máxima do Preview:").pack(anchor="w", padx=5, pady=2)
        self.preview_w_var = ttk.IntVar(value=PREVIEW_W)
        w_spin = ttk.Spinbox(canvas_frame, from_=400, to=1600, increment=100, textvariable=self.preview_w_var, width=10)
        w_spin.pack(anchor="w", padx=5, pady=5)
        
        ttk.Label(canvas_frame, text="Altura Máxima do Preview:").pack(anchor="w", padx=5, pady=(10, 2))
        self.preview_h_var = ttk.IntVar(value=PREVIEW_H)
        h_spin = ttk.Spinbox(canvas_frame, from_=300, to=1200, increment=100, textvariable=self.preview_h_var, width=10)
        h_spin.pack(anchor="w", padx=5, pady=5)
        
        # Configurações Padrão de Detecção
        detection_frame = ttk.LabelFrame(main_frame, text="Configurações Padrão de Detecção", style='Section.TLabelframe')
        detection_frame.pack(fill=X, pady=(0, 10))
        
        ttk.Label(detection_frame, text="Limiar de Correlação Padrão (Clips):").pack(anchor="w", padx=5, pady=2)
        self.thr_corr_var = ttk.DoubleVar(value=THR_CORR)
        corr_frame = ttk.Frame(detection_frame)
        corr_frame.pack(fill=X, padx=5, pady=5)
        
        self.corr_scale = ttk.Scale(corr_frame, from_=0.1, to=1.0, variable=self.thr_corr_var, orient=HORIZONTAL)
        self.corr_scale.pack(side=LEFT, fill=X, expand=True)
        
        self.corr_label = ttk.Label(corr_frame, text=f"{self.thr_corr_var.get():.2f}", width=8)
        self.corr_label.pack(side=RIGHT, padx=(5, 0))
        
        def update_corr_label(val):
            self.corr_label.config(text=f"{float(val):.2f}")
        self.corr_scale.config(command=update_corr_label)
        
        ttk.Label(detection_frame, text="Pixels Mínimos Padrão (Template Matching):").pack(anchor="w", padx=5, pady=(10, 2))
        self.min_px_var = ttk.IntVar(value=MIN_PX)
        px_spin = ttk.Spinbox(detection_frame, from_=1, to=1000, textvariable=self.min_px_var, width=10)
        px_spin.pack(anchor="w", padx=5, pady=5)
        
        # Configurações de Aparência por Local
        appearance_frame = ttk.LabelFrame(main_frame, text="Configurações de Aparência por Local", style='Section.TLabelframe')
        appearance_frame.pack(fill=X, pady=(0, 10))
        
        # Configurações de Fonte para Diferentes Locais
        font_frame = ttk.Frame(appearance_frame)
        font_frame.pack(fill=X, padx=5, pady=5)
        
        # Fonte para Slots
        ttk.Label(font_frame, text="Tamanho da Fonte para Slots:").pack(anchor="w", padx=5, pady=(10, 2))
        self.slot_font_size_var = ttk.IntVar(value=int(self.style_config.get("slot_font_size", 10)))
        slot_font_spin = ttk.Spinbox(font_frame, from_=8, to=24, textvariable=self.slot_font_size_var, width=10)
        slot_font_spin.pack(anchor="w", padx=5, pady=5)
        
        # Fonte para Resultados
        ttk.Label(font_frame, text="Tamanho da Fonte para Resultados:").pack(anchor="w", padx=5, pady=(10, 2))
        self.result_font_size_var = ttk.IntVar(value=int(self.style_config.get("result_font_size", 10)))
        result_font_spin = ttk.Spinbox(font_frame, from_=8, to=24, textvariable=self.result_font_size_var, width=10)
        result_font_spin.pack(anchor="w", padx=5, pady=5)
        
        # Fonte para Botões
        ttk.Label(font_frame, text="Tamanho da Fonte para Botões:").pack(anchor="w", padx=5, pady=(10, 2))
        self.button_font_size_var = ttk.IntVar(value=int(self.style_config.get("button_font_size", 9)))
        button_font_spin = ttk.Spinbox(font_frame, from_=8, to=20, textvariable=self.button_font_size_var, width=10)
        button_font_spin.pack(anchor="w", padx=5, pady=5)
        
        # Configurações de HUD e Inspeção
        hud_frame = ttk.LabelFrame(main_frame, text="Configurações de HUD e Inspeção", style='Section.TLabelframe')
        hud_frame.pack(fill=X, pady=(0, 10))
        
        # Configurações de HUD
        hud_config_frame = ttk.Frame(hud_frame)
        hud_config_frame.pack(fill=X, padx=5, pady=5)
        
        # Tamanho da Fonte do HUD
        ttk.Label(hud_config_frame, text="Tamanho da Fonte do HUD:").pack(anchor="w", padx=5, pady=(10, 2))
        self.hud_font_size_var = ttk.IntVar(value=int(self.style_config.get("hud_font_size", 12)))
        hud_font_spin = ttk.Spinbox(hud_config_frame, from_=8, to=28, textvariable=self.hud_font_size_var, width=10)
        hud_font_spin.pack(anchor="w", padx=5, pady=5)
        
        # Opacidade do HUD
        ttk.Label(hud_config_frame, text="Opacidade do HUD (%):").pack(anchor="w", padx=5, pady=(10, 2))
        self.hud_opacity_var = ttk.IntVar(value=int(self.style_config.get("hud_opacity", 80)))
        opacity_frame = ttk.Frame(hud_config_frame)
        opacity_frame.pack(fill=X, padx=5, pady=5)
        
        self.opacity_scale = ttk.Scale(opacity_frame, from_=10, to=100, variable=self.hud_opacity_var, orient=HORIZONTAL)
        self.opacity_scale.pack(side=LEFT, fill=X, expand=True)
        
        self.opacity_label = ttk.Label(opacity_frame, text=f"{self.hud_opacity_var.get()}%", width=8)
        self.opacity_label.pack(side=RIGHT, padx=(5, 0))
        
        def update_opacity_label(val):
            self.opacity_label.config(text=f"{int(float(val))}%")
        self.opacity_scale.config(command=update_opacity_label)
        
        # Posição do HUD
        ttk.Label(hud_config_frame, text="Posição do HUD:").pack(anchor="w", padx=5, pady=(10, 2))
        self.hud_position_var = ttk.StringVar(value=self.style_config.get("hud_position", "top-right"))
        position_frame = ttk.Frame(hud_config_frame)
        position_frame.pack(fill=X, padx=5, pady=5)
        
        positions = ["top-left", "top-right", "bottom-left", "bottom-right"]
        position_combo = ttk.Combobox(position_frame, textvariable=self.hud_position_var, values=positions, state="readonly", width=15)
        position_combo.pack(side=LEFT, padx=5)
        
        # Mostrar informações adicionais no HUD
        self.show_fps_var = ttk.BooleanVar(value=self.style_config.get("show_fps", True))
        show_fps_check = ttk.Checkbutton(hud_config_frame, text="Mostrar FPS", variable=self.show_fps_var)
        show_fps_check.pack(anchor="w", padx=5, pady=5)
        
        self.show_timestamp_var = ttk.BooleanVar(value=self.style_config.get("show_timestamp", True))
        show_timestamp_check = ttk.Checkbutton(hud_config_frame, text="Mostrar Timestamp", variable=self.show_timestamp_var)
        show_timestamp_check.pack(anchor="w", padx=5, pady=5)
        
        # Cores
        colors_frame = ttk.Frame(appearance_frame)
        colors_frame.pack(fill=X, padx=5, pady=5)
        
        # Cor de Fundo
        bg_color_frame = ttk.Frame(colors_frame)
        bg_color_frame.pack(fill=X, pady=2)
        
        ttk.Label(bg_color_frame, text="Cor de Fundo:").pack(side=LEFT, padx=5)
        self.bg_color_var = ttk.StringVar(value=get_color('colors.background_color', self.style_config))
        bg_color_entry = ttk.Entry(bg_color_frame, textvariable=self.bg_color_var, width=10)
        bg_color_entry.pack(side=LEFT, padx=5)
        
        # Botão para escolher cor
        def choose_bg_color():
            color = colorchooser.askcolor(initialcolor=self.bg_color_var.get(), title="Escolher Cor de Fundo")
            if color and color[1]:
                self.bg_color_var.set(color[1])
        
        ttk.Button(bg_color_frame, text="Escolher", command=choose_bg_color, style='Accent.TButton').pack(side=LEFT, padx=5, pady=2)
        
        # Cor do Texto
        text_color_frame = ttk.Frame(colors_frame)
        text_color_frame.pack(fill=X, pady=2)
        
        ttk.Label(text_color_frame, text="Cor do Texto:").pack(side=LEFT, padx=5)
        self.text_color_var = ttk.StringVar(value=get_color('colors.text_color', self.style_config))
        text_color_entry = ttk.Entry(text_color_frame, textvariable=self.text_color_var, width=10)
        text_color_entry.pack(side=LEFT, padx=5)
        
        # Botão para escolher cor
        def choose_text_color():
            color = colorchooser.askcolor(initialcolor=self.text_color_var.get(), title="Escolher Cor do Texto")
            if color and color[1]:
                self.text_color_var.set(color[1])
        
        ttk.Button(text_color_frame, text="Escolher", command=choose_text_color, style='Accent.TButton').pack(side=LEFT, padx=5, pady=2)
        
        # Cor OK
        ok_color_frame = ttk.Frame(colors_frame)
        ok_color_frame.pack(fill=X, pady=2)
        
        ttk.Label(ok_color_frame, text="Cor OK:").pack(side=LEFT, padx=5)
        self.ok_color_var = ttk.StringVar(value=get_color('colors.ok_color', self.style_config))
        ok_color_entry = ttk.Entry(ok_color_frame, textvariable=self.ok_color_var, width=10)
        ok_color_entry.pack(side=LEFT, padx=5)
        
        # Botão para escolher cor
        def choose_ok_color():
            color = colorchooser.askcolor(initialcolor=self.ok_color_var.get(), title="Escolher Cor OK")
            if color and color[1]:
                self.ok_color_var.set(color[1])
        
        ttk.Button(ok_color_frame, text="Escolher", command=choose_ok_color, style='Success.TButton').pack(side=LEFT, padx=5, pady=2)
        
        # Cor NG
        ng_color_frame = ttk.Frame(colors_frame)
        ng_color_frame.pack(fill=X, pady=2)
        
        ttk.Label(ng_color_frame, text="Cor NG:").pack(side=LEFT, padx=5)
        self.ng_color_var = ttk.StringVar(value=get_color('colors.ng_color', self.style_config))
        ng_color_entry = ttk.Entry(ng_color_frame, textvariable=self.ng_color_var, width=10)
        ng_color_entry.pack(side=LEFT, padx=5)
        
        # Botão para escolher cor
        def choose_ng_color():
            color = colorchooser.askcolor(initialcolor=self.ng_color_var.get(), title="Escolher Cor NG")
            if color and color[1]:
                self.ng_color_var.set(color[1])
        
        ttk.Button(ng_color_frame, text="Escolher", command=choose_ng_color, style='Danger.TButton').pack(side=LEFT, padx=5, pady=2)
        
        # Botões - usando um frame com espaçamento melhor
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=X, pady=(20, 10), padx=10)
        
        # Distribuir os botões uniformemente
        save_btn = ttk.Button(button_frame, text="Salvar", command=self.save_config, style='Success.TButton')
        save_btn.pack(side=LEFT, padx=5, pady=5, expand=True, fill=X)
        
        restore_btn = ttk.Button(button_frame, text="Restaurar Padrões", command=self.restore_defaults, style='Accent.TButton')
        restore_btn.pack(side=LEFT, padx=5, pady=5, expand=True, fill=X)
        
        cancel_btn = ttk.Button(button_frame, text="Cancelar", command=self.cancel, style='Danger.TButton')
        cancel_btn.pack(side=LEFT, padx=5, pady=5, expand=True, fill=X)
    
    def save_config(self):
        """Salva as configurações do sistema"""
        global ORB_FEATURES, ORB_SCALE_FACTOR, ORB_N_LEVELS, PREVIEW_W, PREVIEW_H, THR_CORR, MIN_PX, orb
        
        try:
            # Atualiza variáveis globais
            ORB_FEATURES = int(self.orb_features_var.get())
            ORB_SCALE_FACTOR = float(self.scale_factor_var.get())
            ORB_N_LEVELS = int(self.n_levels_var.get())
            PREVIEW_W = int(self.preview_w_var.get())
            PREVIEW_H = int(self.preview_h_var.get())
            THR_CORR = float(self.thr_corr_var.get())
            MIN_PX = int(self.min_px_var.get())
            
            # Reinicializa detector ORB com novos parâmetros
            try:
                orb = cv2.ORB_create(nfeatures=ORB_FEATURES, scaleFactor=ORB_SCALE_FACTOR, nlevels=ORB_N_LEVELS)
                print(f"Detector ORB reinicializado: features={ORB_FEATURES}, scale={ORB_SCALE_FACTOR}, levels={ORB_N_LEVELS}")
            except Exception as e:
                print(f"Erro ao reinicializar ORB: {e}")
                messagebox.showwarning("Aviso", "Erro ao reinicializar detector ORB. O alinhamento pode não funcionar.")
            
            # Salvar configurações de estilo
            style_config = load_style_config()  # Carrega config atual
            
            # Atualiza configurações de fonte
            style_config["slot_font_size"] = self.slot_font_size_var.get()
            style_config["result_font_size"] = self.result_font_size_var.get()
            style_config["button_font_size"] = self.button_font_size_var.get()
            
            # Atualiza cores na estrutura centralizada
            if "colors" not in style_config:
                style_config["colors"] = {}
            
            style_config["colors"]["background_color"] = self.bg_color_var.get()
            style_config["colors"]["text_color"] = self.text_color_var.get()
            style_config["colors"]["ok_color"] = self.ok_color_var.get()
            style_config["colors"]["ng_color"] = self.ng_color_var.get()
            
            # Salvar configurações de HUD e Inspeção
            style_config["hud_font_size"] = self.hud_font_size_var.get()
            style_config["hud_opacity"] = self.hud_opacity_var.get()
            style_config["hud_position"] = self.hud_position_var.get()
            style_config["show_fps"] = self.show_fps_var.get()
            style_config["show_timestamp"] = self.show_timestamp_var.get()
            
            # Salvar no arquivo de configuração de estilo
            save_style_config(style_config)
            
            # Aplicar as configurações de estilo imediatamente
            apply_style_config(style_config)
            
            self.result = True
            messagebox.showinfo("Sucesso", "Configurações salvas com sucesso!")
            # Desvincular o evento de rolagem do mouse antes de fechar
            try:
                self.unbind_all("<MouseWheel>")
            except:
                pass
            self.destroy()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar configurações: {str(e)}")
    
    def restore_defaults(self):
        """Restaura configurações padrão"""
        # Restaura configurações ORB
        self.orb_features_var.set(5000)
        self.scale_factor_var.set(1.2)
        self.n_levels_var.set(8)
        self.preview_w_var.set(800)
        self.preview_h_var.set(600)
        self.thr_corr_var.set(0.1)
        self.min_px_var.set(10)
        
        # Atualiza labels
        self.features_label.config(text="5000")
        self.scale_label.config(text="1.20")
        self.corr_label.config(text="0.10")
        
        # Restaura configurações de estilo
        self.slot_font_size_var.set(10)
        self.result_font_size_var.set(10)
        self.button_font_size_var.set(9)
        self.bg_color_var.set(get_color('colors.background_color'))
        self.text_color_var.set(get_color('colors.text_color'))
        self.ok_color_var.set(get_color('colors.ok_color'))
        self.ng_color_var.set(get_color('colors.ng_color'))
    
    def cancel(self):
        """Cancela a edição"""
        # Desvincular o evento de rolagem do mouse antes de fechar
        try:
            self.unbind_all("<MouseWheel>")
        except:
            pass
        self.destroy()


class MontagemWindow(ttk.Frame):
    """Wrapper para MontagemWindow no mesh_editor.py"""
    
    def __init__(self, master):
        super().__init__(master)
        try:
            from mesh_editor import MontagemWindow as _MontagemWindow
        except Exception:
            from modulos.mesh_editor import MontagemWindow as _MontagemWindow
        # Importante: instanciar o delegado com 'self' para renderizar dentro desta aba
        self._delegate = _MontagemWindow(self)
        # Renderiza o conteúdo do delegado dentro deste frame
        try:
            self._delegate.pack(fill=BOTH, expand=True)
        except Exception:
            pass
        
        # Delegar atributos importantes
        for attr in ['db_manager', 'img_original', 'slots', 'current_model_id', 'current_model']:
            if hasattr(self._delegate, attr):
                setattr(self, attr, getattr(self._delegate, attr))
    
    def __getattr__(self, name):
        """Delega chamadas de métodos para a instância real."""
        return getattr(self._delegate, name)


class InspecaoWindow(ttk.Frame):
    """Wrapper para InspecaoWindow no inspection_window.py"""
    
    def __init__(self, master):
        super().__init__(master)
        try:
            from inspection_window import InspecaoWindow as _InspecaoWindow
        except Exception:
            from modulos.inspection_window import InspecaoWindow as _InspecaoWindow
        # Importante: instanciar o delegado com 'self' para renderizar dentro desta aba
        self._delegate = _InspecaoWindow(self)
        # Renderiza o conteúdo do delegado dentro deste frame
        try:
            self._delegate.pack(fill=BOTH, expand=True)
        except Exception:
            pass
        
        # Inicializar atributos importantes que outras partes do código esperam
        self.slots = getattr(self._delegate, 'slots', [])
        self.img_reference = getattr(self._delegate, 'img_reference', None)
        self.img_test = getattr(self._delegate, 'img_test', None)
        self.img_display = getattr(self._delegate, 'img_display', None)
        self.scale_factor = getattr(self._delegate, 'scale_factor', 1.0)
        self.current_model_id = getattr(self._delegate, 'current_model_id', None)
        self.inspection_results = getattr(self._delegate, 'inspection_results', [])
        self.camera = getattr(self._delegate, 'camera', None)
        self.live_capture = getattr(self._delegate, 'live_capture', False)
        self.latest_frame = getattr(self._delegate, 'latest_frame', None)
        self.live_view = getattr(self._delegate, 'live_view', False)
        self.available_cameras = getattr(self._delegate, 'available_cameras', [])
        self.selected_camera = getattr(self._delegate, 'selected_camera', 0)
    
    def setup_ui(self):
        """Delega setup_ui para a instância real."""
        return self._delegate.setup_ui()
    
    def update_button_states(self):
        """Delega update_button_states para a instância real."""
        return self._delegate.update_button_states()
    
    def start_live_capture_manual_inspection(self):
        """Delega start_live_capture_manual_inspection para a instância real."""
        return self._delegate.start_live_capture_manual_inspection()
    
    def __getattr__(self, name):
        """Delega chamadas de métodos para a instância real."""
        return getattr(self._delegate, name)


class HistoricoFotosWindow(ttk.Frame):
    def __init__(self, master):
        super().__init__(master)
        
        # Inicializar atributos mínimos usados no setup local
        try:
            self.style = ttk.Style()
        except Exception:
            self.style = None
        from utils import get_color
        try:
            self.accent_color = get_color('colors.ui_colors.primary')
        except Exception:
            self.accent_color = '#6366F1'
        try:
            self.programa_selecionado = ttk.StringVar(value="Todos")
        except Exception:
            self.programa_selecionado = None
        
        # Inicializar banco de dados
        try:
            from database_manager import DatabaseManager
            self.db_manager = DatabaseManager()
        except Exception as e:
            print(f"Erro ao inicializar banco de dados: {e}")
            self.db_manager = None
        
        # Inicializar lista de programas disponíveis
        self.programas_disponiveis = ["Todos"]
        
        # Inicializar listas de fotos (vazias inicialmente)
        self.fotos_historico = []
        self.fotos_ok = []
        self.fotos_ng = []
        self.fotos_capturas = []
        
        # Inicializar listas filtradas
        self.fotos_filtradas = []
        self.fotos_ok_filtradas = []
        self.fotos_ng_filtradas = []
        self.fotos_capturas_filtradas = []
        
        # Modo de exibição: 'Cards' (padrão) ou 'Tabela'
        try:
            self.view_mode = ttk.StringVar(value="Cards")
        except Exception:
            self.view_mode = None
        
        # Diretórios do histórico (garante existência)
        try:
            self._init_history_dirs()
        except Exception:
            pass
        
        # Mapa de normalização de programas (exibição -> normalizado)
        self._program_display_to_norm = {}
        self._program_norm_to_display = {}

    # Utilitários de normalização de nomes de programa
    def _normalize_program(self, name: str) -> str:
        try:
            import unicodedata
            s = (name or '').strip()
            s = unicodedata.normalize('NFKD', s)
            s = ''.join(c for c in s if not unicodedata.combining(c))
            s = s.lower().replace(' ', '_').replace('-', '_')
            while '__' in s:
                s = s.replace('__', '_')
            return s
        except Exception:
            return (name or '').strip().lower().replace(' ', '_')

    def _register_program(self, display_name: str):
        norm = self._normalize_program(display_name)
        self._program_display_to_norm[display_name] = norm
        # Preservar o primeiro display para esse norm
        if norm not in self._program_norm_to_display:
            self._program_norm_to_display[norm] = display_name
        if display_name not in self.programas_disponiveis:
            self.programas_disponiveis.append(display_name)

    def _init_history_dirs(self):
        """Inicializa os diretórios de histórico (ok/ng/capturas)."""
        from paths import get_model_dir
        base_hist = get_model_dir() / 'historico_fotos'
        self.ok_dir = base_hist / 'ok'
        self.ng_dir = base_hist / 'ng'
        self.capturas_dir = base_hist / 'capturas'
        for d in [self.ok_dir, self.ng_dir, self.capturas_dir]:
            d.mkdir(parents=True, exist_ok=True)
        
        # Diretórios do histórico são inicializados por _init_history_dirs no __init__
    
    def setup_ui(self):
        """Configura a interface do usuário."""
        # Limpar widgets existentes para evitar duplicação
        for widget in self.winfo_children():
            widget.destroy()
            
        # Frame principal com layout horizontal (preenche a área do Notebook)
        main_frame = ttk.Frame(self)
        main_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        try:
            # Expansão total do container
            self.pack_propagate(False)
        except Exception:
            pass
        
        # Painel esquerdo - Controles (reduzido ~30%)
        left_panel = ttk.Frame(main_frame)
        # Fixar largura e impedir expansão automática
        left_panel.pack_propagate(False)
        try:
            left_panel.configure(width=238)  # antes 340
        except Exception:
            pass
        left_panel.pack(side=LEFT, fill=Y, padx=(0, 10))
        
        # Cabeçalho com título
        header_frame = ttk.Frame(left_panel)
        header_frame.pack(fill=X, pady=(0, 15))
        
        # Logo e título
        try:
            text_color = get_color('colors.text_color')
            bg_color = get_color('colors.background_color')
        except:
            text_color = '#ffffff'
            bg_color = '#222222'
        header_label = ttk.Label(header_frame, text="DX Project — Histórico de Inspeções", 
                                font=("Segoe UI", 14, "bold"), 
                                foreground=text_color,
                                background=bg_color)
        header_label.pack(pady=10, fill=X)
        
        # Botões de controle
        controls_frame = ttk.LabelFrame(left_panel, text="CONTROLES", style='Section.TLabelframe')
        controls_frame.pack(fill=X, pady=(0, 10))
        
        # Filtro por programa
        filter_frame = ttk.LabelFrame(controls_frame, text="FILTRAR POR PROGRAMA", style='Section.TLabelframe')
        filter_frame.pack(fill=X, padx=5, pady=5)
        
        # Combobox para seleção de programa
        self.programa_combobox = ttk.Combobox(filter_frame, 
                                           textvariable=self.programa_selecionado,
                                           state="readonly")
        self.programa_combobox.pack(fill=X, padx=5, pady=5)
        self.programa_combobox.bind("<<ComboboxSelected>>", self.aplicar_filtros)
        
        # Filtro por período
        period_frame = ttk.LabelFrame(controls_frame, text="FILTRAR POR PERÍODO", style='Section.TLabelframe')
        period_frame.pack(fill=X, padx=5, pady=5)
        
        self.periodo_var = ttk.StringVar(value="30")
        periodo_combobox = ttk.Combobox(period_frame, 
                                       textvariable=self.periodo_var,
                                       values=["7", "30", "90", "365", "Todos"],
                                       state="readonly")
        periodo_combobox.pack(fill=X, padx=5, pady=5)
        periodo_combobox.bind("<<ComboboxSelected>>", self.aplicar_filtros)
        
        try:
            small_font = get_font('small_font')
        except:
            small_font = ('Arial', 7)
        ttk.Label(period_frame, text="dias", font=small_font).pack()
        
        # Modo de exibição (Cards/Tabela)
        view_frame = ttk.LabelFrame(controls_frame, text="MODO DE EXIBIÇÃO", style='Section.TLabelframe')
        view_frame.pack(fill=X, padx=5, pady=(5, 5))
        self.view_mode_combo = ttk.Combobox(view_frame, 
                                            textvariable=self.view_mode,
                                            values=["Cards", "Tabela"],
                                            state="readonly")
        self.view_mode_combo.pack(fill=X, padx=5, pady=5)
        self.view_mode_combo.bind("<<ComboboxSelected>>", self.on_view_mode_change)
        
        # Botão para atualizar histórico
        self.btn_atualizar = ttk.Button(controls_frame, text="ATUALIZAR HISTÓRICO", style='Accent.TButton', 
                                     command=self.atualizar_historico)
        self.btn_atualizar.pack(fill=X, padx=5, pady=5)
        
        # Botão para limpar histórico
        self.btn_limpar = ttk.Button(controls_frame, text="LIMPAR HISTÓRICO", style='Danger.TButton', 
                                   command=self.limpar_historico)
        self.btn_limpar.pack(fill=X, padx=5, pady=5)

        # Botão para exportar relatório em Excel
        self.btn_exportar = ttk.Button(controls_frame, text="EXPORTAR RELATÓRIO (Excel)", style='Accent.TButton',
                                       command=self.exportar_relatorio_excel)
        self.btn_exportar.pack(fill=X, padx=5, pady=5)
        
        # Painel direito - Histórico de fotos
        right_panel = ttk.Frame(main_frame)
        right_panel.pack(side=RIGHT, fill=BOTH, expand=True)
        # Garantir expansão total usando pack; grid não é necessário aqui
        # Expandir notebook para ocupar toda a área disponível
        
        # Notebook para organizar fotos por categoria
        self.historico_notebook = ttk.Notebook(right_panel)
        self.historico_notebook.pack(fill=BOTH, expand=True)
        
        # Aba para todas as fotos
        self.todas_fotos_frame = ttk.Frame(self.historico_notebook)
        self.historico_notebook.add(self.todas_fotos_frame, text="Todas as Fotos")
        
        # Aba para fotos OK
        self.ok_fotos_frame = ttk.Frame(self.historico_notebook)
        self.historico_notebook.add(self.ok_fotos_frame, text="Aprovadas (OK)")
        
        # Aba para fotos NG
        self.ng_fotos_frame = ttk.Frame(self.historico_notebook)
        self.historico_notebook.add(self.ng_fotos_frame, text="Reprovadas (NG)")
        
        # Aba para capturas manuais
        self.capturas_fotos_frame = ttk.Frame(self.historico_notebook)
        self.historico_notebook.add(self.capturas_fotos_frame, text="Capturas Manuais")
        
        # Aba para estatísticas
        self.estatisticas_frame = ttk.Frame(self.historico_notebook)
        self.historico_notebook.add(self.estatisticas_frame, text="Estatísticas")
        
        # Criar scrollable frames para cada aba (sempre expandindo para preencher)
        self.todas_scrollable_frame = self.criar_scrollable_frame(self.todas_fotos_frame, "todas")
        self.ok_scrollable_frame = self.criar_scrollable_frame(self.ok_fotos_frame, "ok")
        self.ng_scrollable_frame = self.criar_scrollable_frame(self.ng_fotos_frame, "ng")
        self.capturas_scrollable_frame = self.criar_scrollable_frame(self.capturas_fotos_frame, "capturas")
        
        # Configurar aba de estatísticas
        self.setup_estatisticas_ui()
        
        # Não carregar dados automaticamente - apenas carregar programas disponíveis
        self.carregar_programas_disponiveis()

        # Renderização sob demanda por aba
        try:
            self.historico_notebook.bind("<<NotebookTabChanged>>", self.on_hist_tab_changed)
        except Exception:
            pass
        
        # Mostrar mensagem inicial
        self.mostrar_mensagem_inicial()
    
    def setup_estatisticas_ui(self):
        """Configura a interface de estatísticas."""
        # Frame principal para estatísticas
        stats_main_frame = ttk.Frame(self.estatisticas_frame)
        stats_main_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        # Estatísticas gerais
        general_stats_frame = ttk.LabelFrame(stats_main_frame, text="ESTATÍSTICAS GERAIS", style='Section.TLabelframe')
        general_stats_frame.pack(fill=X, pady=(0, 10))
        
        self.stats_labels = {}
        stats_grid = ttk.Frame(general_stats_frame)
        stats_grid.pack(fill=X, padx=10, pady=10)
        
        # Criar grid de estatísticas (2 colunas)
        stats_items = [
            ("Total de Inspeções:", "total_inspections"),
            ("Taxa de Aprovação:", "approval_rate"),
            ("Tempo Médio (s):", "avg_processing_time"),
            ("Confiança Média:", "avg_confidence")
        ]
        
        for i, (label_text, key) in enumerate(stats_items):
            row = i // 2
            col = i % 2
            
            # Label do nome da estatística
            try:
                font = get_font('small_font')
            except:
                font = ('Arial', 7)
            label = ttk.Label(stats_grid, text=label_text, font=font)
            label.grid(row=row, column=col*2, sticky="w", padx=(10 if col == 0 else 20), pady=5)
            
            # Label do valor da estatística
            try:
                color = get_color('colors.ui_colors.primary')
            except:
                color = '#6366F1'
            value_label = ttk.Label(stats_grid, text="0", font=font, foreground=color)
            value_label.grid(row=row, column=col*2+1, sticky="w", padx=5, pady=5)
            
            self.stats_labels[key] = value_label
        
        # Resumo por modelo
        model_summary_frame = ttk.LabelFrame(stats_main_frame, text="RESUMO POR MODELO", style='Section.TLabelframe')
        model_summary_frame.pack(fill=BOTH, expand=True, pady=(0, 10))
        
        # Treeview para resumo por modelo
        columns = ("Modelo", "Total", "OK", "NG", "Taxa OK", "Última Inspeção")
        self.model_tree = ttk.Treeview(model_summary_frame, columns=columns, show="headings", height=8)
        
        # Configurar colunas
        for col in columns:
            self.model_tree.heading(col, text=col)
            self.model_tree.column(col, width=120, anchor="center")
        
        # Scrollbar para a treeview
        model_scrollbar = ttk.Scrollbar(model_summary_frame, orient="vertical", command=self.model_tree.yview)
        self.model_tree.configure(yscrollcommand=model_scrollbar.set)
        
        self.model_tree.pack(side=LEFT, fill=BOTH, expand=True, padx=(10, 0), pady=10)
        model_scrollbar.pack(side=RIGHT, fill=Y, pady=10)

    def exportar_relatorio_excel(self):
        """Exporta o histórico e o resumo para um arquivo Excel (.xlsx)."""
        try:
            import csv
            try:
                import openpyxl  # noqa: F401
            except Exception:
                pass  # Se não houver openpyxl, o pandas/csv simples ainda pode ser usado

            from tkinter import filedialog, messagebox
            from database_manager import DatabaseManager
            db = self.db_manager or DatabaseManager()

            # Filtros atuais
            modelo_nome = None if (self.programa_selecionado is None or self.programa_selecionado.get() == 'Todos') else self.programa_selecionado.get()
            try:
                dias = 0 if self.periodo_var.get() == 'Todos' else int(self.periodo_var.get())
            except Exception:
                dias = 30

            # Buscar dados
            historico = db.get_inspection_history(modelo_nome=modelo_nome, limit=100000, offset=0)
            estat = db.get_inspection_statistics(modelo_nome=modelo_nome, days=(dias or 36500))
            resumo = db.get_model_inspection_summary()

            # Selecionar destino
            default_name = "relatorio_inspecoes.xlsx"
            file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=default_name,
                                                     filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv"), ("Todos", "*.*")])
            if not file_path:
                return

            if file_path.lower().endswith('.csv'):
                # Exporta histórico principal em CSV
                keys = historico[0].keys() if historico else ['id','modelo_id','modelo_nome','slot_id','result','confidence','processing_time','image_path','created_at']
                with open(file_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=keys)
                    writer.writeheader()
                    for row in historico:
                        writer.writerow(row)
            else:
                # Exporta em Excel com múltiplas abas
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
                wb = Workbook()
                ws_hist = wb.active
                ws_hist.title = 'Historico'
                # Cabeçalhos
                headers = list(historico[0].keys()) if historico else ['id','modelo_id','modelo_nome','slot_id','result','confidence','processing_time','image_path','created_at']
                ws_hist.append(headers)
                for row in historico:
                    ws_hist.append([row.get(k, '') for k in headers])
                # Ajuste de largura
                for i, h in enumerate(headers, start=1):
                    ws_hist.column_dimensions[get_column_letter(i)].width = min(max(12, len(str(h)) + 2), 40)

                # Aba Estatísticas
                ws_est = wb.create_sheet('Estatisticas')
                for k in ['total_inspections','total_ok','total_ng','approval_rate','avg_processing_time','avg_confidence','period_days']:
                    ws_est.append([k, estat.get(k, '')])

                # Aba Resumo por Modelo
                ws_sum = wb.create_sheet('Resumo_Modelos')
                if resumo:
                    headers_sum = list(resumo[0].keys())
                    ws_sum.append(headers_sum)
                    for row in resumo:
                        ws_sum.append([row.get(k, '') for k in headers_sum])
                    for i, h in enumerate(headers_sum, start=1):
                        ws_sum.column_dimensions[get_column_letter(i)].width = min(max(12, len(str(h)) + 2), 40)

                wb.save(file_path)

            messagebox.showinfo("Relatório", f"Relatório exportado com sucesso:\n{file_path}")
        except Exception as e:
            try:
                from tkinter import messagebox
                messagebox.showerror("Erro", f"Falha ao exportar relatório: {e}")
            except Exception:
                print(f"Falha ao exportar relatório: {e}")
    
    def carregar_programas_disponiveis(self):
        """Carrega apenas a lista de programas disponíveis sem carregar fotos."""
        try:
            if self.db_manager:
                # Obter lista de modelos disponíveis
                modelos = self.db_manager.list_modelos()
                self.programas_disponiveis = ["Todos"]
                
                for modelo in modelos:
                    if modelo['nome'] not in self.programas_disponiveis:
                        self.programas_disponiveis.append(modelo['nome'])
                
                # Atualizar combobox de programas
                if hasattr(self, 'programa_combobox'):
                    self.programa_combobox['values'] = self.programas_disponiveis
                    self.programa_combobox.current(0)  # Selecionar "Todos"
                    
        except Exception as e:
            print(f"Erro ao carregar programas disponíveis: {e}")
    
    def mostrar_mensagem_inicial(self):
        """Mostra mensagem inicial indicando que o usuário deve clicar em ATUALIZAR."""
        try:
            # Mostrar mensagem em todas as abas
            for frame_name in ['todas', 'ok', 'ng', 'capturas']:
                frame = getattr(self, f"{frame_name}_scrollable_frame")
                if frame:
                    # Limpar frame
                    for widget in frame.winfo_children():
                        widget.destroy()
                    
                    # Criar mensagem centralizada
                    msg_frame = ttk.Frame(frame)
                    msg_frame.pack(expand=True, fill=BOTH)
                    
                    try:
                        font = get_font('subtitle_font')
                        color = get_color('colors.special_colors.gray_text')
                    except:
                        font = ('Arial', 16)
                        color = '#888888'
                    
                    ttk.Label(msg_frame, 
                             text="📸 Histórico de Inspeções", 
                             font=font, 
                             foreground=color).pack(pady=(50, 20))
                    
                    try:
                        small_font = get_font('small_font')
                    except:
                        small_font = ('Arial', 10)
                    
                    ttk.Label(msg_frame, 
                             text="Clique em 'ATUALIZAR HISTÓRICO' para carregar os dados", 
                             font=small_font, 
                             foreground=color).pack(pady=10)
                    
                    ttk.Label(msg_frame, 
                             text="Use os filtros acima para refinar sua busca", 
                             font=small_font, 
                             foreground=color).pack()
            
            # Mostrar estatísticas vazias
            self.exibir_estatisticas()
            
        except Exception as e:
            print(f"Erro ao mostrar mensagem inicial: {e}")
    
    def mostrar_mensagem_sem_fotos(self):
        """Mostra mensagem quando não há fotos carregadas."""
        try:
            # Mostrar mensagem em todas as abas
            for frame_name in ['todas', 'ok', 'ng', 'capturas']:
                frame = getattr(self, f"{frame_name}_scrollable_frame")
                if frame:
                    # Limpar frame
                    for widget in frame.winfo_children():
                        widget.destroy()
                    
                    # Criar mensagem centralizada
                    msg_frame = ttk.Frame(frame)
                    msg_frame.pack(expand=True, fill=BOTH)
                    
                    try:
                        font = get_font('subtitle_font')
                        color = get_color('colors.special_colors.gray_text')
                    except:
                        font = ('Arial', 16)
                        color = '#888888'
                    
                    ttk.Label(msg_frame, 
                             text="📸 Nenhuma Foto Carregada", 
                             font=font, 
                             foreground=color).pack(pady=(50, 20))
                    
                    try:
                        small_font = get_font('small_font')
                    except:
                        small_font = ('Arial', 10)
                    
                    ttk.Label(msg_frame, 
                             text="Clique em 'ATUALIZAR HISTÓRICO' para carregar os dados", 
                             font=small_font, 
                             foreground=color).pack(pady=10)
                    
                    ttk.Label(msg_frame, 
                             text="Use os filtros acima para refinar sua busca", 
                             font=small_font, 
                             foreground=color).pack()
            
        except Exception as e:
            print(f"Erro ao mostrar mensagem sem fotos: {e}")
    
    def carregar_dados_iniciais(self):
        """Carrega dados iniciais do histórico."""
        try:
            # Carregar fotos existentes (legado)
            self.carregar_fotos_existentes()
            
            # Carregar dados do banco de dados
            if self.db_manager:
                self.carregar_dados_banco()
            
            # Exibir fotos
            self.exibir_fotos()
            
            # Exibir estatísticas
            self.exibir_estatisticas()
            
        except Exception as e:
            print(f"Erro ao carregar dados iniciais: {e}")
    
    def carregar_dados_banco(self):
        """Carrega dados do banco de dados."""
        try:
            if not self.db_manager:
                return
            
            # Obter lista de modelos disponíveis
            modelos = self.db_manager.list_modelos()
            self.programas_disponiveis = ["Todos"]
            
            for modelo in modelos:
                if modelo['nome'] not in self.programas_disponiveis:
                    self.programas_disponiveis.append(modelo['nome'])
            
            # Atualizar combobox de programas
            if hasattr(self, 'programa_combobox'):
                self.programa_combobox['values'] = self.programas_disponiveis
                self.programa_combobox.current(0)  # Selecionar "Todos"
                
        except Exception as e:
            print(f"Erro ao carregar dados do banco: {e}")
    
    def carregar_fotos_existentes(self):
        """Carrega as fotos existentes no diretório de histórico com suporte a versões otimizadas."""
        try:
            # Limpar listas existentes
            self.fotos_historico = []
            self.fotos_ok = []
            self.fotos_ng = []
            self.fotos_capturas = []
            
            # Função auxiliar para processar arquivos de uma pasta
            def processar_arquivos(diretorio, categoria):
                fotos = []
                if diretorio.exists():
                    for arquivo in diretorio.glob("*.*"):
                        # Considerar apenas imagens
                        if arquivo.suffix.lower() not in ['.png', '.jpg', '.jpeg']:
                            continue
                        stem = arquivo.stem
                        # Base sem sufixos de otimização
                        if stem.endswith('_hist'):
                            base_stem = stem[:-5]
                        elif stem.endswith('_thumb'):
                            base_stem = stem[:-6]
                        else:
                            base_stem = stem

                        try:
                            nome = arquivo.name
                            nome_sem_ext = base_stem  # sem extensão e sem sufixos
                            timestamp = None
                            programa = "Desconhecido"

                            # Remove prefixos conhecidos e tokens supérfluos no final
                            base = nome_sem_ext
                            if categoria == "capturas" and base.startswith("foto_"):
                                base = base[5:]
                            if (categoria in ("ok", "ng")) and base.startswith("inspecao_"):
                                base = base[9:]

                            # Tokens por '_', removendo sufixos 'hist'/'thumb' caso tenham passado
                            partes = [p for p in base.split('_') if p]
                            if partes and partes[-1].lower() in {"hist", "thumb"}:
                                partes = partes[:-1]

                            # Procura par (YYYYMMDD, HHMMSS)
                            def eh_data(s):
                                return len(s) == 8 and s.isdigit()
                            def eh_hora(s):
                                return len(s) == 6 and s.isdigit()

                            idx_data = None
                            for i in range(max(0, len(partes) - 6), len(partes) - 1):
                                if eh_data(partes[i]) and eh_hora(partes[i+1]):
                                    idx_data = i
                                    break

                            if idx_data is not None:
                                data_str = partes[idx_data]
                                hora_str = partes[idx_data + 1]
                                try:
                                    timestamp = datetime.strptime(f"{data_str}_{hora_str}", "%Y%m%d_%H%M%S")
                                except Exception:
                                    timestamp = None
                                programa = "_".join(partes[:idx_data]) if idx_data > 0 else "Desconhecido"

                            # Fallback: usa mtime do arquivo
                            if timestamp is None:
                                try:
                                    timestamp = datetime.fromtimestamp(arquivo.stat().st_mtime)
                                except Exception:
                                    pass

                            if timestamp is None:
                                # Não conseguiu timestamp, ignora silenciosamente
                                continue

                            # Seleciona versão otimizada para exibição se existir (usando base)
                            arquivo_hist = arquivo.parent / f"{base_stem}_hist.jpg"
                            arquivo_thumb = arquivo.parent / f"{base_stem}_thumb.jpg"
                            # Escolhe original se existir; senão usa o arquivo atual
                            original_png = arquivo.parent / f"{base_stem}.png"
                            original_jpg = arquivo.parent / f"{base_stem}.jpg"
                            original_jpeg = arquivo.parent / f"{base_stem}.jpeg"
                            arquivo_original = original_png if original_png.exists() else (original_jpg if original_jpg.exists() else (original_jpeg if original_jpeg.exists() else arquivo))
                            arquivo_exibicao = arquivo_hist if arquivo_hist.exists() else (arquivo_thumb if arquivo_thumb.exists() else arquivo_original)

                            # Registrar programa (display) e normalizado
                            if programa and programa != "Desconhecido":
                                self._register_program(programa.replace('_', ' '))

                            fotos.append({
                                'arquivo': arquivo_original,
                                'arquivo_exibicao': arquivo_exibicao,
                                'timestamp': timestamp,
                                'categoria': categoria,
                                'programa': programa.replace('_', ' ') if programa else 'Desconhecido',
                                'programa_norm': self._normalize_program(programa),
                                'tem_otimizada': arquivo_hist.exists() or arquivo_thumb.exists()
                            })
                        except Exception as e:
                            print(f"Erro ao processar arquivo {arquivo}: {e}")
                return fotos
            
            # Processar arquivos de cada diretório (padrão)
            self.fotos_ok = processar_arquivos(self.ok_dir, "ok")
            self.fotos_ng = processar_arquivos(self.ng_dir, "ng")
            self.fotos_capturas = processar_arquivos(self.capturas_dir, "capturas")
            
            # Combinar todas as fotos
            self.fotos_historico = self.fotos_ok + self.fotos_ng + self.fotos_capturas
            
            # Fallback 1: se nada encontrado, tentar varrer subpastas diretas de historico_fotos
            if not self.fotos_historico:
                base_hist = self.ok_dir.parent  # historico_fotos
                if base_hist.exists():
                    for sub in base_hist.iterdir():
                        if not sub.is_dir():
                            continue
                        if sub.name.lower() in {"ok", "ng", "capturas"}:
                            continue
                        fotos_sub = processar_arquivos(sub, "capturas")
                        # Se programa estiver desconhecido, usar nome da pasta
                        if fotos_sub:
                            for f in fotos_sub:
                                if f.get('programa', 'Desconhecido') == 'Desconhecido':
                                    f['programa'] = sub.name
                                    f['programa_norm'] = self._normalize_program(sub.name)
                                    self._register_program(sub.name)
                            self.fotos_capturas.extend(fotos_sub)
                    self.fotos_historico = self.fotos_ok + self.fotos_ng + self.fotos_capturas

            # Fallback 2: se ainda nada, procurar imagens soltas na raiz de historico_fotos
            if not self.fotos_historico:
                base_hist = self.ok_dir.parent
                if base_hist.exists():
                    fotos_root = []
                    for arquivo in base_hist.glob("*.*"):
                        if arquivo.suffix.lower() not in ['.png', '.jpg', '.jpeg']:
                            continue
                        # Reutilizar processamento via diretório raiz
                        try:
                            stem = arquivo.stem
                            base_stem = stem[:-5] if stem.endswith('_hist') else (stem[:-6] if stem.endswith('_thumb') else stem)
                            # Timestamp via mtime
                            timestamp = datetime.fromtimestamp(arquivo.stat().st_mtime)
                            # Arquivo exibição
                            arquivo_hist = arquivo.parent / f"{base_stem}_hist.jpg"
                            arquivo_thumb = arquivo.parent / f"{base_stem}_thumb.jpg"
                            original_png = arquivo.parent / f"{base_stem}.png"
                            original_jpg = arquivo.parent / f"{base_stem}.jpg"
                            original_jpeg = arquivo.parent / f"{base_stem}.jpeg"
                            arquivo_original = original_png if original_png.exists() else (original_jpg if original_jpg.exists() else (original_jpeg if original_jpeg.exists() else arquivo))
                            arquivo_exibicao = arquivo_hist if arquivo_hist.exists() else (arquivo_thumb if arquivo_thumb.exists() else arquivo_original)
                            fotos_root.append({
                                'arquivo': arquivo_original,
                                'arquivo_exibicao': arquivo_exibicao,
                                'timestamp': timestamp,
                                'categoria': 'capturas',
                                'programa': 'Desconhecido',
                                'programa_norm': self._normalize_program('Desconhecido'),
                                'tem_otimizada': arquivo_hist.exists() or arquivo_thumb.exists()
                            })
                        except Exception:
                            continue
                    if fotos_root:
                        self.fotos_capturas.extend(fotos_root)
                        self.fotos_historico = self.fotos_ok + self.fotos_ng + self.fotos_capturas

            # Ordenar por timestamp (mais recente primeiro)
            self.fotos_historico.sort(key=lambda x: x['timestamp'], reverse=True)
            self.fotos_ok.sort(key=lambda x: x['timestamp'], reverse=True)
            self.fotos_ng.sort(key=lambda x: x['timestamp'], reverse=True)
            self.fotos_capturas.sort(key=lambda x: x['timestamp'], reverse=True)
            
        except Exception as e:
            print(f"Erro ao carregar fotos existentes: {e}")
    
    def exibir_fotos(self):
        """Exibe as fotos no histórico."""
        try:
            # Verificar se a interface foi inicializada
            if not hasattr(self, "todas_scrollable_frame") or self.todas_scrollable_frame is None:
                print("Interface não inicializada completamente. Tentando inicializar...")
                self.setup_ui()
                # Carregar fotos existentes
                self.carregar_fotos_existentes()
                # Atualizar combobox de programas
                if self.programa_combobox is not None:
                    self.programa_combobox['values'] = self.programas_disponiveis
                    self.programa_combobox.current(0)  # Selecionar "Todos"
                return
                
            # Verificar se há fotos carregadas
            if not hasattr(self, 'fotos_historico') or not self.fotos_historico:
                print("Nenhuma foto carregada. Clique em ATUALIZAR HISTÓRICO primeiro.")
                # Mostrar mensagem nas abas
                self.mostrar_mensagem_sem_fotos()
                return
            
            # Usar listas filtradas se disponíveis, senão listas originais
            if hasattr(self, 'fotos_filtradas') and self.fotos_filtradas:
                fotos_por_cat = {
                    'todas': self.fotos_filtradas,
                    'ok': self.fotos_ok_filtradas,
                    'ng': self.fotos_ng_filtradas,
                    'capturas': self.fotos_capturas_filtradas,
                }
            else:
                fotos_por_cat = {
                    'todas': self.fotos_historico,
                    'ok': self.fotos_ok,
                    'ng': self.fotos_ng,
                    'capturas': self.fotos_capturas,
                }

            # Renderiza somente a aba atual para economizar recursos
            try:
                current_tab = self.historico_notebook.tab(self.historico_notebook.select(), 'text')
            except Exception:
                current_tab = 'Todas as Fotos'

            aba_map = {
                'Todas as Fotos': ('todas', self.todas_scrollable_frame),
                'Aprovadas (OK)': ('ok', self.ok_scrollable_frame),
                'Reprovadas (NG)': ('ng', self.ng_scrollable_frame),
                'Capturas Manuais': ('capturas', self.capturas_scrollable_frame),
            }

            cat, frame = aba_map.get(current_tab, ('todas', self.todas_scrollable_frame))
            fotos = fotos_por_cat.get(cat, [])
            print(f"Exibindo {len(fotos)} fotos na aba '{current_tab}' - modo {self.view_mode.get() if self.view_mode else 'Cards'}")
            if self.view_mode and self.view_mode.get() == "Tabela":
                self.exibir_fotos_em_tabela(frame, fotos, cat)
            else:
                self.exibir_fotos_em_aba(frame, fotos, cat)
        except Exception as e:
            print(f"Erro ao exibir fotos: {e}")

    def on_hist_tab_changed(self, event=None):
        """Re-renderiza a aba de histórico selecionada ao trocar de aba."""
        try:
            # Reutiliza listas já carregadas/filtradas
            if hasattr(self, 'fotos_filtradas') and self.fotos_filtradas:
                fotos_por_cat = {
                    'todas': self.fotos_filtradas,
                    'ok': self.fotos_ok_filtradas,
                    'ng': self.fotos_ng_filtradas,
                    'capturas': self.fotos_capturas_filtradas,
                }
            else:
                fotos_por_cat = {
                    'todas': self.fotos_historico,
                    'ok': self.fotos_ok,
                    'ng': self.fotos_ng,
                    'capturas': self.fotos_capturas,
                }

            current_tab = self.historico_notebook.tab(self.historico_notebook.select(), 'text')
            aba_map = {
                'Todas as Fotos': ('todas', self.todas_scrollable_frame),
                'Aprovadas (OK)': ('ok', self.ok_scrollable_frame),
                'Reprovadas (NG)': ('ng', self.ng_scrollable_frame),
                'Capturas Manuais': ('capturas', self.capturas_scrollable_frame),
            }
            cat, frame = aba_map.get(current_tab, ('todas', self.todas_scrollable_frame))
            fotos = fotos_por_cat.get(cat, [])
            if self.view_mode and self.view_mode.get() == "Tabela":
                self.exibir_fotos_em_tabela(frame, fotos, cat)
            else:
                self.exibir_fotos_em_aba(frame, fotos, cat)
        except Exception as e:
            print(f"Erro ao alternar aba do histórico: {e}")
    
    def exibir_estatisticas(self):
        """Exibe as estatísticas das inspeções."""
        try:
            if not self.db_manager:
                return
            
            # Verificar se há fotos carregadas
            if not hasattr(self, 'fotos_historico') or not self.fotos_historico:
                # Mostrar estatísticas vazias
                self.stats_labels['total_inspections'].config(text="0")
                self.stats_labels['approval_rate'].config(text="0%")
                self.stats_labels['avg_processing_time'].config(text="0.0")
                self.stats_labels['avg_confidence'].config(text="0.0")
                
                # Limpar treeview
                for item in self.model_tree.get_children():
                    self.model_tree.delete(item)
                return
            
            # Obter programa selecionado (normalizado)
            programa_display = self.programa_selecionado.get()
            programa_norm_sel = self._normalize_program(programa_display) if programa_display != 'Todos' else 'todos'
            periodo = self.periodo_var.get()
            
            # Calcular estatísticas baseadas nas fotos carregadas
            if hasattr(self, 'fotos_filtradas') and self.fotos_filtradas:
                # Usar fotos filtradas
                fotos_para_estatisticas = self.fotos_filtradas
                print(f"Calculando estatísticas para {len(fotos_para_estatisticas)} fotos filtradas")
            else:
                # Usar todas as fotos carregadas
                fotos_para_estatisticas = self.fotos_historico
                print(f"Calculando estatísticas para {len(fotos_para_estatisticas)} fotos totais")
            
            # Calcular estatísticas manualmente
            total_inspections = len(fotos_para_estatisticas)
            ok_count = len([f for f in fotos_para_estatisticas if f.get('categoria') == 'ok'])
            ng_count = len([f for f in fotos_para_estatisticas if f.get('categoria') == 'ng'])
            
            approval_rate = round((ok_count / total_inspections * 100) if total_inspections > 0 else 0, 1)
            
            # Calcular tempo médio e confiança média (se disponíveis)
            avg_processing_time = "N/A"
            avg_confidence = "N/A"
            
            # Atualizar labels de estatísticas
            self.stats_labels['total_inspections'].config(text=str(total_inspections))
            self.stats_labels['approval_rate'].config(text=f"{approval_rate}%")
            self.stats_labels['avg_processing_time'].config(text=str(avg_processing_time))
            self.stats_labels['avg_confidence'].config(text=str(avg_confidence))
            
            # Calcular resumo por modelo
            modelos_stats = {}
            for foto in fotos_para_estatisticas:
                programa_foto = foto.get('programa', 'Desconhecido')
                if programa_foto not in modelos_stats:
                    modelos_stats[programa_foto] = {'total': 0, 'ok': 0, 'ng': 0}
                
                modelos_stats[programa_foto]['total'] += 1
                if foto.get('categoria') == 'ok':
                    modelos_stats[programa_foto]['ok'] += 1
                elif foto.get('categoria') == 'ng':
                    modelos_stats[programa_foto]['ng'] += 1
            
            # Limpar treeview
            for item in self.model_tree.get_children():
                self.model_tree.delete(item)
            
            # Preencher treeview com estatísticas calculadas
            for modelo_nome, stats in modelos_stats.items():
                if programa_norm_sel == 'todos' or self._normalize_program(modelo_nome) == programa_norm_sel:
                    approval_rate_modelo = round((stats['ok'] / stats['total'] * 100) if stats['total'] > 0 else 0, 1)
                    self.model_tree.insert("", "end", values=(
                        modelo_nome,
                        stats['total'],
                        stats['ok'],
                        stats['ng'],
                        f"{approval_rate_modelo}%",
                        "N/A"  # Última inspeção não disponível nas fotos
                    ))
                    
        except Exception as e:
            print(f"Erro ao exibir estatísticas: {e}")
    
    def aplicar_filtros(self, event=None):
        """Aplica filtros por programa e período às fotos já carregadas."""
        try:
            programa = self.programa_selecionado.get()
            periodo = self.periodo_var.get()
            
            print(f"Filtros aplicados - Programa: {programa}, Período: {periodo} dias")
            
            # Verificar se o programa está na lista de programas disponíveis
            if programa not in self.programas_disponiveis:
                print(f"Programa {programa} não encontrado na lista de programas disponíveis")
                return
            
            # Verificar se há fotos carregadas
            if not hasattr(self, 'fotos_historico') or not self.fotos_historico:
                print("Nenhuma foto carregada. Clique em ATUALIZAR HISTÓRICO primeiro.")
                return
                
            # Aplicar filtros
            self.aplicar_filtros_por_programa_e_data(programa, periodo)
            
            # Atualizar exibição de fotos
            self.exibir_fotos()
            
            # Atualizar estatísticas
            self.exibir_estatisticas()
            
        except Exception as e:
            print(f"Erro ao aplicar filtros: {e}")
    
    def aplicar_filtros_por_programa_e_data(self, programa, periodo):
        """Aplica filtros por programa e data às fotos já carregadas."""
        try:
            from datetime import datetime, timedelta
            
            # Calcular data limite se não for "Todos"
            if periodo != "Todos":
                dias = int(periodo)
                data_limite = datetime.now() - timedelta(days=dias)
            else:
                data_limite = None
            
            # Filtrar fotos por programa e data
            fotos_filtradas = []
            
            for foto in self.fotos_historico:
                # Filtrar por programa (normalizado)
                if programa != "Todos":
                    pf = foto.get('programa_norm') or self._normalize_program(foto.get('programa', ''))
                    if pf != self._normalize_program(programa):
                        continue
                    continue
                
                # Filtrar por data se aplicável
                if data_limite and 'timestamp' in foto:
                    if foto['timestamp'] < data_limite:
                        continue
                
                fotos_filtradas.append(foto)
            
            # Atualizar listas filtradas
            self.fotos_filtradas = fotos_filtradas
            
            # Separar por categoria
            self.fotos_ok_filtradas = [f for f in fotos_filtradas if f.get('categoria') == 'ok']
            self.fotos_ng_filtradas = [f for f in fotos_filtradas if f.get('categoria') == 'ng']
            self.fotos_capturas_filtradas = [f for f in fotos_filtradas if f.get('categoria') == 'capturas']
            
            # Garantir que as listas filtradas existam
            if not hasattr(self, 'fotos_filtradas'):
                self.fotos_filtradas = []
            if not hasattr(self, 'fotos_ok_filtradas'):
                self.fotos_ok_filtradas = []
            if not hasattr(self, 'fotos_ng_filtradas'):
                self.fotos_ng_filtradas = []
            if not hasattr(self, 'fotos_capturas_filtradas'):
                self.fotos_capturas_filtradas = []
            
            print(f"Filtros aplicados: {len(fotos_filtradas)} fotos de {len(self.fotos_historico)} total")
            
        except Exception as e:
            print(f"Erro ao aplicar filtros por programa e data: {e}")
    
    def filtrar_por_programa(self, event=None):
        """Método legado - redireciona para aplicar_filtros."""
        self.aplicar_filtros(event)
    
    def atualizar_historico(self):
        """Atualiza a exibição do histórico de fotos."""
        try:
            # Carregar dados do banco de dados
            if self.db_manager:
                self.carregar_dados_banco()
            
            # Carregar fotos existentes
            self.carregar_fotos_existentes()
            
            # Aplicar filtros atuais se houver
            programa = self.programa_selecionado.get()
            periodo = self.periodo_var.get()
            
            if programa != "Todos" or periodo != "30":
                # Aplicar filtros atuais
                self.aplicar_filtros_por_programa_e_data(programa, periodo)
                print(f"Filtros aplicados após atualização: Programa={programa}, Período={periodo}")
            
            # Exibir fotos
            self.exibir_fotos()
            
            # Exibir estatísticas
            self.exibir_estatisticas()
            
            messagebox.showinfo("Sucesso", f"Histórico atualizado!\n{len(self.fotos_historico)} fotos carregadas")
        except Exception as e:
            print(f"Erro ao atualizar histórico: {e}")
            messagebox.showerror("Erro", f"Erro ao atualizar histórico: {e}")
    
    def limpar_historico(self):
        """Limpa o histórico de fotos."""
        try:
            if messagebox.askyesno("Confirmar", "Deseja realmente limpar todo o histórico de fotos?"):
                # Limpar arquivos (incluindo versões otimizadas)
                for diretorio in [self.ok_dir, self.ng_dir, self.capturas_dir]:
                    if diretorio.exists():
                        # Limpar arquivos PNG originais e suas versões otimizadas
                        for arquivo in diretorio.glob("*.png"):
                            try:
                                # Deletar arquivo original
                                arquivo.unlink()
                                
                                # Deletar versões otimizadas se existirem
                                base_name = arquivo.stem
                                hist_file = diretorio / f"{base_name}_hist.jpg"
                                thumb_file = diretorio / f"{base_name}_thumb.jpg"
                                
                                if hist_file.exists():
                                    hist_file.unlink()
                                if thumb_file.exists():
                                    thumb_file.unlink()
                                    
                            except Exception as e:
                                print(f"Erro ao excluir arquivo {arquivo}: {e}")
                        
                        # Limpar arquivos JPG otimizados restantes
                        for arquivo in diretorio.glob("*_hist.jpg"):
                            try:
                                arquivo.unlink()
                            except Exception as e:
                                print(f"Erro ao excluir arquivo otimizado {arquivo}: {e}")
                        
                        for arquivo in diretorio.glob("*_thumb.jpg"):
                            try:
                                arquivo.unlink()
                            except Exception as e:
                                print(f"Erro ao excluir thumbnail {arquivo}: {e}")
                
                # Limpar listas
                self.fotos_historico = []
                self.fotos_ok = []
                self.fotos_ng = []
                self.fotos_capturas = []
                
                # Limpar estatísticas
                if self.db_manager:
                    self.db_manager.clear_inspection_history()
                
                # Atualizar interface
                self.exibir_fotos()
                self.exibir_estatisticas()
                
                messagebox.showinfo("Sucesso", "Histórico limpo com sucesso!\nTodas as versões otimizadas também foram removidas.")
        except Exception as e:
            print(f"Erro ao limpar histórico: {e}")
            messagebox.showerror("Erro", f"Erro ao limpar histórico: {e}")
    
    def criar_scrollable_frame(self, parent_frame, categoria):
        """Cria um frame com scrollbar para exibir fotos."""
        try:
            # Limpar widgets existentes para evitar duplicação
            for widget in parent_frame.winfo_children():
                widget.destroy()
                
            try:
                bg_color = get_color('colors.canvas_colors.canvas_bg')
            except:
                bg_color = '#1E1E1E'
            canvas = Canvas(parent_frame, bg=bg_color, highlightthickness=0)
            scrollbar = ttk.Scrollbar(parent_frame, orient="vertical", command=canvas.yview)
            scrollable_frame = ttk.Frame(canvas)
            
            scrollable_frame.bind(
                "<Configure>",
                lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
            )
            
            window_id = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)
            
            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")
            
            # Adiciona suporte para scroll com mouse wheel
            canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
            # Garante que o conteúdo ocupe a largura do canvas
            def _sync_width(event=None, wid=window_id):
                try:
                    canvas.itemconfig(wid, width=canvas.winfo_width())
                except Exception:
                    pass
            canvas.bind("<Configure>", _sync_width)
            _sync_width()
            
            # Armazenar referências
            setattr(self, f"{categoria}_canvas", canvas)
            setattr(self, f"{categoria}_scrollbar", scrollbar)
            setattr(self, f"{categoria}_scrollable_frame", scrollable_frame)
            setattr(self, f"{categoria}_window_id", window_id)
            
            return scrollable_frame
        except Exception as e:
            print(f"Erro ao criar frame scrollable para {categoria}: {e}")
            return None
    
    def exibir_fotos_em_aba(self, frame, fotos, categoria):
        """Exibe as fotos em uma aba específica."""
        # Verificar se o frame existe
        if frame is None:
            print(f"Frame para categoria {categoria} não foi inicializado corretamente")
            return
        
        # Limpar frame existente
        for widget in frame.winfo_children():
            widget.destroy()
        
        if not fotos:
            # Mensagem quando não há fotos
            try:
                font = get_font('subtitle_font')
                color = get_color('colors.special_colors.gray_text')
            except:
                font = ('Arial', 24)
                color = '#888888'
            ttk.Label(frame, 
                     text="Nenhuma foto nesta categoria", 
                     font=font, 
                     foreground=color).pack(pady=20)
            return
        
        # Layout em grade (responsivo) com 3 colunas
        grid_frame = ttk.Frame(frame)
        grid_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)

        num_cols = 3
        for c in range(num_cols):
            grid_frame.grid_columnconfigure(c, weight=1)

        row, col = 0, 0
        # Limitar máximo em tela para evitar esgotar recursos do Windows (GDI/menus)
        max_cards = 300
        for i, foto_info in enumerate(fotos[:max_cards]):
            try:
                if 'arquivo' in foto_info and foto_info['arquivo'].exists():
                    card = self.criar_card_foto(grid_frame, foto_info)
                    if card is not None:
                        card.grid(row=row, column=col, sticky='nsew', padx=5, pady=8)
                    col += 1
                    if col >= num_cols:
                        col = 0
                        row += 1
                else:
                    print(f"Arquivo não encontrado para foto {i} na categoria {categoria}")
            except Exception as e:
                print(f"Erro ao criar card para foto {foto_info.get('arquivo', 'desconhecida')}: {e}")
                continue
    
    def criar_card_foto(self, parent_frame, foto_info):
        """Cria um card para exibir uma foto com suas informações."""
        try:
            # Frame para o card
            card_frame = ttk.Frame(parent_frame, relief="solid", borderwidth=1)
            # A posição é definida pelo caller (grid)
            
            # Usar arquivo de exibição otimizado se disponível
            arquivo_exibicao = foto_info.get('arquivo_exibicao', foto_info['arquivo'])
            
            # Carregar e exibir a imagem
            img = cv2.imread(str(arquivo_exibicao))
            if img is not None:
                # Redimensionar para exibição (thumbnail) e manter referência para não coletar
                img_height, img_width = img.shape[:2]
                max_width = 320
                scale = 1.0 if img_width <= max_width else (max_width / float(img_width))
                display_width = int(img_width * scale)
                display_height = int(img_height * scale)
                img_resized = img if scale == 1.0 else cv2.resize(img, (display_width, display_height))
                
                # Converter para formato Tkinter
                img_rgb = cv2.cvtColor(img_resized, cv2.COLOR_BGR2RGB)
                img_pil = Image.fromarray(img_rgb)
                img_tk = ImageTk.PhotoImage(img_pil)
                
                # Label para a imagem
                img_label = ttk.Label(card_frame, image=img_tk)
                # Guardar referência forte no frame para evitar GC do Tk
                if not hasattr(card_frame, '_images'):
                    card_frame._images = []
                card_frame._images.append(img_tk)
                img_label.pack(pady=5)
                
                # Informações da foto
                info_frame = ttk.Frame(card_frame)
                info_frame.pack(fill=X, padx=10, pady=5)
                
                # Data e hora
                timestamp = foto_info['timestamp']
                data_str = timestamp.strftime("%d/%m/%Y")
                hora_str = timestamp.strftime("%H:%M:%S")
                
                # Categoria e programa
                categoria = foto_info.get('categoria', 'desconhecida')
                programa = foto_info.get('programa', 'Desconhecido')
                
                # Cor baseada na categoria
                try:
                    categoria_cor = get_color('colors.status_colors.success_bg') if categoria == "ok" else \
                               get_color('colors.status_colors.error_bg') if categoria == "ng" else \
                               get_color('colors.status_colors.info_bg') if categoria == "capturas" else get_color('colors.status_colors.neutral_bg')
                except:
                    categoria_cor = '#00AA00' if categoria == "ok" else \
                               '#CC0000' if categoria == "ng" else \
                               '#0066CC' if categoria == "capturas" else '#888888'
                
                categoria_texto = "APROVADO" if categoria == "ok" else \
                                 "REPROVADO" if categoria == "ng" else \
                                 "CAPTURA MANUAL" if categoria == "capturas" else "DESCONHECIDO"
                
                try:
                    font_small = get_font('small_font')
                    font_tiny = get_font('tiny_font')
                except:
                    font_small = ('Arial', 7)
                    font_tiny = ('Arial', 8)
                
                ttk.Label(info_frame, text=f"📊 Status: {categoria_texto}", 
                         font=font_small, foreground=categoria_cor).pack(anchor="w")
                ttk.Label(info_frame, text=f"🔧 Programa: {programa}", font=font_small).pack(anchor="w")
                ttk.Label(info_frame, text=f"📅 Data: {data_str}", font=font_small).pack(anchor="w")
                ttk.Label(info_frame, text=f"🕒 Hora: {hora_str}", font=font_small).pack(anchor="w")
                ttk.Label(info_frame, text=f"📏 Dimensões: {display_width}x{display_height}", font=font_tiny).pack(anchor="w")
                
                # Indicador de otimização
                if foto_info.get('tem_otimizada', False):
                    try:
                        color = get_color('colors.ui_colors.success')
                    except:
                        color = '#10B981'
                    ttk.Label(info_frame, text="⚡ Otimizada", font=font_tiny, foreground=color).pack(anchor="w")
                
                # Botões de ação
                btn_frame = ttk.Frame(card_frame)
                btn_frame.pack(fill=X, padx=10, pady=5)
                
                # Botão para visualizar em tamanho real
                btn_visualizar = ttk.Button(btn_frame, text="Visualizar", 
                                         command=lambda fi=foto_info: self.visualizar_foto(fi))
                btn_visualizar.pack(side=LEFT, padx=5)
                
                # Botão para excluir
                btn_excluir = ttk.Button(btn_frame, text="Excluir", 
                                       command=lambda fi=foto_info, cf=card_frame: self.excluir_foto(fi, cf))
                btn_excluir.pack(side=RIGHT, padx=5)
            else:
                ttk.Label(card_frame, text="Erro ao carregar imagem", foreground="red").pack(pady=10)
            
            return card_frame
        
        except Exception as e:
            print(f"Erro ao criar card para foto: {e}")
            return None

    def exibir_fotos_em_tabela(self, frame, fotos, categoria):
        """Exibe as fotos em uma tabela (Treeview) rápida de carregar."""
        try:
            # Limpar frame existente
            for widget in frame.winfo_children():
                widget.destroy()
            
            columns = ("status", "programa", "data", "hora", "dimensoes", "arquivo")
            tv = ttk.Treeview(frame, columns=columns, show="headings", height=20)
            # Cabeçalhos
            tv.heading("status", text="Status")
            tv.heading("programa", text="Programa")
            tv.heading("data", text="Data")
            tv.heading("hora", text="Hora")
            tv.heading("dimensoes", text="Dimensões")
            tv.heading("arquivo", text="Arquivo")
            # Larguras
            tv.column("status", width=100, anchor="center")
            tv.column("programa", width=160, anchor="w")
            tv.column("data", width=100, anchor="center")
            tv.column("hora", width=90, anchor="center")
            tv.column("dimensoes", width=110, anchor="center")
            tv.column("arquivo", width=360, anchor="w")

            # Scrollbar vertical
            vsb = ttk.Scrollbar(frame, orient="vertical", command=tv.yview)
            tv.configure(yscrollcommand=vsb.set)
            tv.pack(side=LEFT, fill=BOTH, expand=True)
            vsb.pack(side=RIGHT, fill=Y)

            # Popular linhas (sem imagem para máxima performance)
            for f in fotos:
                try:
                    ts = f.get('timestamp')
                    data_str = ts.strftime("%d/%m/%Y") if ts else ""
                    hora_str = ts.strftime("%H:%M:%S") if ts else ""
                    cat = f.get('categoria', '')
                    status = "APROVADO" if cat == 'ok' else ("REPROVADO" if cat == 'ng' else "CAPTURA")
                    dims = ""
                    try:
                        # Se o arquivo de exibição existe, lê dimensões sem carregar tudo
                        img_path = f.get('arquivo_exibicao') or f.get('arquivo')
                        if img_path and img_path.exists():
                            # Evitar abrir a imagem inteira; mas para simplicidade, mostramos string padrão
                            dims = "320x240"
                    except Exception:
                        pass
                    tv.insert("", "end", values=(status,
                                                    f.get('programa', ''),
                                                    data_str,
                                                    hora_str,
                                                    dims,
                                                    str((f.get('arquivo') or '')).replace('\\', '/')))
                except Exception:
                    continue

            # Eventos: duplo-clique para visualizar, Delete para excluir
            def on_open_selected(event=None):
                try:
                    sel = tv.selection()
                    if not sel:
                        return
                    idx = tv.index(sel[0])
                    if 0 <= idx < len(fotos):
                        self.visualizar_foto(fotos[idx])
                except Exception:
                    pass
            def on_delete_selected(event=None):
                try:
                    sel = tv.selection()
                    if not sel:
                        return
                    idx = tv.index(sel[0])
                    if 0 <= idx < len(fotos):
                        # Precisa encontrar o frame/card; aqui reusa excluir com None e re-renderiza
                        self.excluir_foto(fotos[idx], card_frame=None)
                except Exception:
                    pass
            tv.bind("<Double-1>", on_open_selected)
            tv.bind("<Delete>", on_delete_selected)
        except Exception as e:
            print(f"Erro ao exibir tabela: {e}")

    def on_view_mode_change(self, event=None):
        try:
            self.exibir_fotos()
        except Exception:
            pass
    
    def visualizar_foto(self, foto_info):
        """Abre uma janela para visualizar a foto em tamanho real com zoom."""
        try:
            # Usar arquivo de exibição otimizado se disponível, mas para visualização
            # preferir a imagem original para máxima qualidade
            arquivo_visualizacao = foto_info['arquivo']  # Sempre usar original para visualização
            
            img = cv2.imread(str(arquivo_visualizacao))
            if img is not None:
                # Criar janela de visualização
                # Cria janela filha do root da aplicação para evitar erro de menu/contexto
                try:
                    root_window = self.winfo_toplevel()
                except Exception:
                    root_window = None
                view_window = Toplevel(master=root_window or self)
                view_window.title(f"Foto - {foto_info['timestamp'].strftime('%d/%m/%Y %H:%M:%S')}")
                # Evitar janela modal/transient que possa conflitar com menus do ttkbootstrap
                
                # Ajustar tamanho da janela (máximo 80% da tela)
                screen_width = view_window.winfo_screenwidth()
                screen_height = view_window.winfo_screenheight()
                
                img_height, img_width = img.shape[:2]
                scale = min(0.8 * screen_width / img_width, 0.8 * screen_height / img_height)
                
                if scale < 1:  # Redimensionar apenas se for maior que 80% da tela
                    new_width = int(img_width * scale)
                    new_height = int(img_height * scale)
                    img = cv2.resize(img, (new_width, new_height))
                
                # Converter para formato Tkinter (evitar imagens muito grandes)
                img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                max_dim = 4000
                h0, w0 = img_rgb.shape[:2]
                if max(h0, w0) > max_dim:
                    scale0 = max_dim / float(max(h0, w0))
                    img_rgb = cv2.resize(img_rgb, (int(w0*scale0), int(h0*scale0)))
                img_pil = Image.fromarray(img_rgb)
                img_tk = ImageTk.PhotoImage(img_pil)
                
                # Canvas para exibir a imagem com scrollbars
                canvas_frame = ttk.Frame(view_window)
                canvas_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
                
                h_scrollbar = ttk.Scrollbar(canvas_frame, orient=HORIZONTAL)
                h_scrollbar.pack(side=BOTTOM, fill=X)
                
                v_scrollbar = ttk.Scrollbar(canvas_frame, orient=VERTICAL)
                v_scrollbar.pack(side=RIGHT, fill=Y)
                
                canvas = Canvas(canvas_frame, 
                               xscrollcommand=h_scrollbar.set,
                               yscrollcommand=v_scrollbar.set)
                canvas.pack(side=LEFT, fill=BOTH, expand=True)
                
                h_scrollbar.config(command=canvas.xview)
                v_scrollbar.config(command=canvas.yview)
                
                # Exibir imagem no canvas e manter IDs
                img_item = canvas.create_image(0, 0, anchor=NW, image=img_tk)
                canvas.image = img_tk  # Manter referência forte

                # Configurar região de rolagem baseada no tamanho da imagem
                canvas.config(scrollregion=(0, 0, img_tk.width(), img_tk.height()))

                # Estado de zoom/pan local para esta janela de visualização
                zoom_state = {
                    'level': 1.0,
                    'orig_img': img_rgb,
                    'img_item': img_item,
                    'img_w0': img_width,
                    'img_h0': img_height,
                }

                def resize_to_level(level: float, anchor_old_x: float = None, anchor_old_y: float = None):
                    # Limita o zoom
                    level = max(0.2, min(level, 3.0))
                    old_w = int(zoom_state['img_w0'] * zoom_state['level'])
                    old_h = int(zoom_state['img_h0'] * zoom_state['level'])
                    new_w = int(zoom_state['img_w0'] * level)
                    new_h = int(zoom_state['img_h0'] * level)

                    # Redimensiona imagem
                    img_resized = cv2.resize(zoom_state['orig_img'], (new_w, new_h))
                    img_pil2 = Image.fromarray(img_resized)
                    tk2 = ImageTk.PhotoImage(img_pil2)
                    canvas.itemconfig(zoom_state['img_item'], image=tk2)
                    canvas.image = tk2  # Atualiza referência

                    # Atualiza scrollregion para novo tamanho
                    canvas.config(scrollregion=(0, 0, new_w, new_h))

                    # Mantém o ponto sob o cursor ancorado após o zoom
                    try:
                        if anchor_old_x is not None and anchor_old_y is not None and old_w > 0 and old_h > 0:
                            rat_x = max(0.0, min(1.0, anchor_old_x / float(old_w)))
                            rat_y = max(0.0, min(1.0, anchor_old_y / float(old_h)))
                            new_x = rat_x * new_w
                            new_y = rat_y * new_h
                            # Ajusta a visão para centralizar próximo ao ponto
                            vx = max(0.0, min(1.0, (new_x - canvas.winfo_width() / 2) / max(new_w, 1)))
                            vy = max(0.0, min(1.0, (new_y - canvas.winfo_height() / 2) / max(new_h, 1)))
                            canvas.xview_moveto(vx)
                            canvas.yview_moveto(vy)
                    except Exception:
                        pass

                    zoom_state['level'] = level

                # Função para aplicar zoom com foco no cursor
                def apply_zoom(event):
                    try:
                        delta = int(event.delta)
                    except Exception:
                        delta = 0

                    # Ponto atual sob o cursor em coordenadas de conteúdo
                    old_w = int(zoom_state['img_w0'] * zoom_state['level'])
                    old_h = int(zoom_state['img_h0'] * zoom_state['level'])
                    anchor_x = canvas.canvasx(event.x)
                    anchor_y = canvas.canvasy(event.y)

                    # Passo de zoom suave (~10%)
                    if delta > 0:
                        new_level = zoom_state['level'] * 1.1
                    elif delta < 0:
                        new_level = zoom_state['level'] / 1.1
                    else:
                        new_level = zoom_state['level']

                    resize_to_level(new_level, anchor_x, anchor_y)

                # Pan/arrasto usando scan_mark/scan_dragto (suave) com botão esquerdo e direito
                def start_pan(event):
                    try:
                        canvas.scan_mark(event.x, event.y)
                        canvas.config(cursor="fleur")
                    except Exception:
                        pass

                def do_pan(event):
                    try:
                        canvas.scan_dragto(event.x, event.y, gain=1)
                    except Exception:
                        pass

                def end_pan(event):
                    try:
                        canvas.config(cursor="")
                    except Exception:
                        pass

                # Bindings
                try:
                    canvas.bind("<MouseWheel>", apply_zoom, add=True)
                except Exception:
                    canvas.bind("<MouseWheel>", apply_zoom)
                canvas.bind("<ButtonPress-1>", start_pan)
                canvas.bind("<B1-Motion>", do_pan)
                canvas.bind("<ButtonRelease-1>", end_pan)
                canvas.bind("<ButtonPress-3>", start_pan)
                canvas.bind("<B3-Motion>", do_pan)
                canvas.bind("<ButtonRelease-3>", end_pan)
                # Duplo clique para resetar
                def reset_zoom(event=None):
                    resize_to_level(1.0, 0, 0)
                    canvas.xview_moveto(0.0)
                    canvas.yview_moveto(0.0)
                canvas.bind("<Double-1>", reset_zoom)
                
                # Controles inferiores
                controls = ttk.Frame(view_window)
                controls.pack(fill=X, pady=10)
                ttk.Button(controls, text="Fechar", command=view_window.destroy).pack(side=RIGHT, padx=10)
                ttk.Button(controls, text="Redefinir Zoom", command=lambda: resize_to_level(1.0, 0, 0)).pack(side=RIGHT)
                try:
                    view_window.lift()
                    view_window.focus_force()
                except Exception:
                    pass
            else:
                messagebox.showerror("Erro", "Não foi possível carregar a imagem.")
        except Exception as e:
            print(f"Erro ao visualizar foto: {e}")
            messagebox.showerror("Erro", f"Erro ao visualizar foto: {e}")
    
    def excluir_foto(self, foto_info, card_frame):
        """Exclui uma foto do histórico e suas versões otimizadas."""
        try:
            if messagebox.askyesno("Confirmar", "Deseja realmente excluir esta foto do histórico?"):
                # Excluir arquivo original
                if foto_info['arquivo'].exists():
                    foto_info['arquivo'].unlink()
                
                # Excluir versões otimizadas se existirem
                arquivo_base = foto_info['arquivo']
                arquivo_hist = arquivo_base.parent / f"{arquivo_base.stem}_hist.jpg"
                arquivo_thumb = arquivo_base.parent / f"{arquivo_base.stem}_thumb.jpg"
                
                if arquivo_hist.exists():
                    arquivo_hist.unlink()
                    print(f"Versão de histórico excluída: {arquivo_hist}")
                
                if arquivo_thumb.exists():
                    arquivo_thumb.unlink()
                    print(f"Thumbnail excluído: {arquivo_thumb}")
                
                # Remover da lista
                self.fotos_historico = [f for f in self.fotos_historico if f['arquivo'] != foto_info['arquivo']]
                cat = foto_info.get('categoria')
                if cat == 'ok' and hasattr(self, 'fotos_ok'):
                    self.fotos_ok = [f for f in self.fotos_ok if f['arquivo'] != foto_info['arquivo']]
                elif cat == 'ng' and hasattr(self, 'fotos_ng'):
                    self.fotos_ng = [f for f in self.fotos_ng if f['arquivo'] != foto_info['arquivo']]
                elif cat == 'capturas' and hasattr(self, 'fotos_capturas'):
                    self.fotos_capturas = [f for f in self.fotos_capturas if f['arquivo'] != foto_info['arquivo']]
                # Atualiza listas filtradas, se existirem
                if hasattr(self, 'fotos_filtradas') and self.fotos_filtradas:
                    self.fotos_filtradas = [f for f in self.fotos_filtradas if f['arquivo'] != foto_info['arquivo']]
                if hasattr(self, 'fotos_ok_filtradas') and self.fotos_ok_filtradas:
                    self.fotos_ok_filtradas = [f for f in self.fotos_ok_filtradas if f['arquivo'] != foto_info['arquivo']]
                if hasattr(self, 'fotos_ng_filtradas') and self.fotos_ng_filtradas:
                    self.fotos_ng_filtradas = [f for f in self.fotos_ng_filtradas if f['arquivo'] != foto_info['arquivo']]
                if hasattr(self, 'fotos_capturas_filtradas') and self.fotos_capturas_filtradas:
                    self.fotos_capturas_filtradas = [f for f in self.fotos_capturas_filtradas if f['arquivo'] != foto_info['arquivo']]
                
                # Remover card da interface
                card_frame.destroy()
                
                messagebox.showinfo("Sucesso", "Foto e versões otimizadas excluídas com sucesso!")
                # Re-renderiza aba atual para evitar lacunas
                try:
                    self.exibir_fotos()
                    self.exibir_estatisticas()
                except Exception:
                    pass
        except Exception as e:
            print(f"Erro ao excluir foto: {e}")
            messagebox.showerror("Erro", f"Erro ao excluir foto: {e}")
    
    def start_background_frame_capture(self):
        """Inicia a captura contínua de frames em segundo plano."""
        def capture_frames():
            while self.live_capture and self.camera and self.camera.isOpened():
                try:
                    ret, frame = self.camera.read()
                    if ret:
                        self.latest_frame = frame.copy()
                    time.sleep(0.033)  # ~30 FPS
                except Exception as e:
                    print(f"Erro na captura de frame: {e}")
                    break
        
        import threading
        self.capture_thread = threading.Thread(target=capture_frames, daemon=True)
        self.capture_thread.start()


def apply_dynamic_scaling(root, target_width: int = 1920, target_height: int = 1080,
                          min_scale: float = 0.9, max_scale: float = 1.1) -> None:
    """Aplica escala global de UI proporcional à resolução atual.

    - Usa como referência 1920x1080 (Full HD) para manter proporções em telas menores
    - Ajusta 'tk scaling' (afeta todos os tamanhos de fonte em pontos)
    - Ajusta fontes nomeadas padrão do Tk para coerência geral
    """
    try:
        screen_w = max(1, int(root.winfo_screenwidth()))
        screen_h = max(1, int(root.winfo_screenheight()))

        # Fator baseado na menor razão de largura/altura
        proportional_scale = min(screen_w / float(target_width), screen_h / float(target_height))
        # Mantém a escala dentro de limites conservadores para não reduzir demais a legibilidade
        proportional_scale = max(min_scale, min(proportional_scale, max_scale))

        # Ajuste do 'tk scaling' mantendo o DPI base atual
        try:
            current_scaling = float(root.tk.call('tk', 'scaling'))
        except Exception:
            current_scaling = 1.0
        # Não reduz mais do que 10% do scaling atual e nunca abaixo do scaling atual do Tk
        # (muitos ambientes já vêm com scaling de DPI configurado pelo SO)
        target_scaling = current_scaling * float(proportional_scale)
        min_allowed = current_scaling * 0.9
        new_scaling = max(min_allowed, target_scaling)
        root.tk.call('tk', 'scaling', new_scaling)

        # Ajusta fontes nomeadas padrão (reflete em diversos widgets ttk)
        try:
            import tkinter.font as tkfont
            named_fonts = [
                "TkDefaultFont", "TkTextFont", "TkFixedFont", "TkMenuFont",
                "TkHeadingFont", "TkCaptionFont", "TkSmallCaptionFont",
                "TkIconFont", "TkTooltipFont"
            ]
            for font_name in named_fonts:
                try:
                    f = tkfont.nametofont(font_name)
                    base_size = int(f.cget("size"))
                    # Redução limitada a 10% e aumento moderado
                    scaled_size = int(round(base_size * proportional_scale))
                    min_size = max(6, int(round(base_size * 0.9)))
                    max_size = int(round(base_size * 1.1))
                    new_size = max(min_size, min(scaled_size, max_size))
                    if new_size != base_size:
                        f.configure(size=new_size)
                except Exception:
                    pass
        except Exception:
            pass
    except Exception:
        # Não bloqueia startup em caso de inconsistência de DPI
        pass


def create_main_window():
    """Cria e configura a janela principal da aplicação com funcionalidades avançadas."""
    import ttkbootstrap as ttk
    from ttkbootstrap import Style
    
    # Inicializa ttkbootstrap com tema moderno
    root = ttk.Window(
        title="DX Project — Sistema de Inspeção Visual",
        themename="cyborg",  # Tema moderno escuro mais profissional
        size=(1400, 900),
        resizable=(True, True)
    )
    # Aplica escala proporcional antes de construir o restante da UI
    apply_dynamic_scaling(root)
    
    # Configurar para abrir maximizada no Windows
    try:
        root.state('zoomed')  # Maximiza a janela no Windows
    except:
        # Fallback para outros sistemas
        root.place_window_center()
    
    # Configurar ícone da janela (se disponível)
    try:
        root.iconbitmap(str(get_project_root() / "assets" / "dx_project_logo.png"))
    except:
        try:
            # Fallback para outros formatos de ícone
            root.iconphoto(True, tk.PhotoImage(file=str(get_project_root() / "assets" / "dx_project_logo.png")))
        except:
            pass  # Ignora se não encontrar o ícone
    
    # Aplica tema/estilo centralizado
    try:
        style_cfg = load_style_config()
        apply_style_config(style_cfg)
    except Exception:
        pass
    
    # Configura fechamento de janelas OpenCV e limpeza de recursos
    def on_closing():
        cv2.destroyAllWindows()
        # Limpa cache de câmeras antes de fechar
        try:
            release_all_cached_cameras()
            print("Cache de câmeras limpo ao fechar aplicação principal")
        except Exception as e:
            print(f"Erro ao limpar cache de câmeras na aplicação principal: {e}")
        root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_closing)

    # Cabeçalho profissional com título e estilo
    header = ttk.Frame(root, padding=(15, 10))
    header.pack(fill="x")
    title_lbl = ttk.Label(
        header, 
                    text="DX Project — Sistema de Inspeção Visual", 
        style="TLabel",
        font=("Segoe UI", 16, "bold")
    )
    title_lbl.pack(side=LEFT)

    # Criar notebook para as abas com estilo moderno
    notebook = ttk.Notebook(root)
    notebook.pack(fill="both", expand=True, padx=15, pady=15)
    
    # Criar as abas com funcionalidades avançadas
    # Aba Editor de Malha
    montagem_frame = MontagemWindow(notebook)
    notebook.add(montagem_frame, text="📝 Editor de Malha")
    
    # Aba Inspeção
    inspecao_frame = InspecaoWindow(notebook)
    notebook.add(inspecao_frame, text="🔍 Inspeção em Tempo Real")
    
    # Aba Histórico
    historico_frame = HistoricoFotosWindow(notebook)
    # Garante que a UI do histórico seja construída
    if hasattr(historico_frame, "setup_ui"):
        historico_frame.setup_ui()
    notebook.add(historico_frame, text="📊 Histórico e Estatísticas")

    # Garantir que a aba "Editor de Malha" seja selecionada por padrão
    notebook.select(0)
    
    # Adicionar evento para detectar mudança de aba com funcionalidades automáticas
    def on_tab_changed(event):
        # Verificar se a aba selecionada é a de Inspeção (índice 1)
        if notebook.index(notebook.select()) == 1:
            # Iniciar captura da câmera automaticamente quando mudar para aba de inspeção
            try:
                if hasattr(inspecao_frame, "start_live_capture_manual_inspection"):
                    inspecao_frame.start_live_capture_manual_inspection()
            except Exception as e:
                print(f"Erro ao iniciar captura automática: {e}")
    
    # Vincular evento de mudança de aba
    notebook.bind("<<NotebookTabChanged>>", on_tab_changed)
    
    return root


def main():
    """Função principal do módulo montagem."""
    root = create_main_window()
    root.mainloop()
    return root


if __name__ == "__main__":
    main()